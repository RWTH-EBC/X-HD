import dash
import dash_core_components as dcc
import dash_html_components as html
import Functions_GUI.Functions as Func
import Functions_GUI.Design_functions as desFunc
from dash.dependencies import Input, Output
import openpyxl as xl
import time
import plotly.express as px
import pandas as pd
import os
from pathlib import Path

# path to this module
dirpath = os.path.dirname(__file__)
# path to default config
config_default_path = 'Functions_GUI/config/default/config_default.json'
# get config file
data_default = Func.open_config(config_default_path)
# path to config
config_path = "Functions_GUI/config/config.json"
# write new config
Func.write_config(data_default, config_path)
# open new config
data = Func.open_config(config_path)
# absolut path
abs_config_path = os.path.join(dirpath, 'Functions_GUI\\config\\config.json')

columns = ['demand_heat', 'demand_dhw', 'produced_heat', 'electrical_hp', 'electrical_hr', 'cop', 'amb_temp',
           'sto_dhw_top', 'sto_buf_top', 'time']
df_res_din = pd.DataFrame(columns=columns)
df_res_vdi = pd.DataFrame(columns=columns)
df_res_ind = pd.DataFrame(columns=columns)

# error to check, if all necessary data was given
fill_error = True

# counter for button clicks
click_counter = 0

# define a color set
colors = {
    'black': '#111111',
    'white': '#FFFFFF',
    'ebcred': '#DD402D',
    'ebcgrey': '#9D9EA0'
}

# define default style for html boxes
def_style = {'borderLeft': '1px solid ' + str(colors['black']), 'padding': 10}
box_style = {'padding': 15, 'backgroundColor': colors['ebcgrey'], 'marginBottom': '2em'}
input_style = {'color': colors['black'], 'borderBottom': '1px solid ' + str(colors['black']), 'marginBottom': '1em'}

external_stylesheets = ['https://codepen.io/chriddyp/pen/bWLwgP.css']

height = 200
height_two = 200

app = dash.Dash(__name__, external_stylesheets=external_stylesheets)


app.layout = html.Div(style={'backgroundColor': colors['white']}, children=[
    html.Img(src='https://www.ebc.eonerc.rwth-aachen.de/global/show_picture.asp?id=aaaaaaaaaakevlz',
             style={'height': 90}
             ),

    html.H1(
        children='Normative HPS-Design',
        style={
            'textAlign': 'center',
            'color': colors['black'],
            'display': 'inline-block-center'
        }
    ),

    html.Div(children='A web application for the normative design of heatpump systems', style={
        'textAlign': 'center',
        'marginBottom': '3em'}
             ),

    html.Div(children=[

        dcc.Tabs(id="tabs", value='tab-calc', children=[
            dcc.Tab(label='Calculation specifications', value='tab-calc', children=[
                html.Div(children=[
                    dcc.Markdown(style={'color': colors['black'],
                                        'border': '1px solid ' + str(colors['black']),
                                        'marginBottom': '1em',
                                        'marginTop': '1em',
                                        'padding': 15},
                                 children='''
                        #### Inputs for calculation process
                        Here you can choose the calculations performed by the tool. For the calculation of the heat 
                        demand over a year, select "Heat Demand". Further information regarding the ambient temperature 
                        and the building type are needed. For the calculation of the normative Designs according to the 
                        DIN and VDI standards, select the corresponding one.
                        Furthermore the calculation of the SCOP of the normative HPS System and full-year simulations of
                        the system are possible.

                        To determine the calculations with Teaser, the inputs from the tab "Building data"
                        are required.
                        '''),
                ]),
                html.Div(className="row", style={'padding': 15, 'backgroundColor': colors['ebcred'],
                                                 'marginBottom': '2em'},
                         children=[
                             html.Div(className="four columns",
                                      style={'padding': 10},
                                      children=[
                                          dcc.Markdown(style={'color': colors['black'],
                                                              'borderBottom': '1px solid ' + str(colors['black']),
                                                              'marginBottom': '1em'},
                                                       children='''Which calculation shall be performed?'''),

                                          dcc.Checklist(id='check-calc', options=[
                                                  {'label': 'Heat demand (Teaser)', 'value': 'heat-demand'},
                                                  {'label': 'DIN', 'value': 'DIN'},
                                                  {'label': 'VDI', 'value': 'VDI'}
                                              ],
                                              value=['DIN', 'VDI']
                                          ),
                                      ]),

                             html.Div(id='weather-dat', className="four columns",
                                      style={'borderLeft': '1px solid ' + str(colors['black']), 'padding': 10,
                                             'display': 'none'},
                                      children=[
                                          dcc.Markdown(style={'color': colors['black'],
                                                              'borderBottom': '1px solid ' + str(colors['black']),
                                                              'marginBottom': '1em'},
                                                       children='''Data for weather'''),

                                          dcc.RadioItems(id='check-weather', options=[
                                                  {'label': 'TRY', 'value': 'try'},
                                                  {'label': 'Individual temperature', 'value': 'ind'}
                                              ],
                                              value='try'
                                          )
                                      ]),
                             html.Div(id='ind-weather', className="four columns",
                                      style={'borderLeft': '1px solid ' + str(colors['black']), 'padding': 10,
                                             'display': 'none'},
                                      children=[
                                          dcc.Markdown(style={'color': colors['black'],
                                                              'borderBottom': '1px solid ' + str(colors['black']),
                                                              'marginBottom': '1em'},
                                                       children='''Desired ambient temperature in °C'''),

                                          dcc.Input(id='ind-weatherval',
                                                    placeholder='Enter a value...',
                                                    type='text',
                                                    value=None)
                                      ]),
                         ]),

                html.Div(className="row", style={'padding': 15, 'backgroundColor': colors['ebcgrey'],
                                                 'marginBottom': '2em'},
                         children=[
                             html.Div(className="four columns",
                                      style={'padding': 10},
                                      children=[
                                          html.Div(style={'height': height_two}, children=[
                                              dcc.Markdown(style={'color': colors['black'],
                                                                  'borderBottom': '1px solid ' + str(
                                                                      colors['black']),
                                                                  'marginBottom': '1em'},
                                                           children='''Annual Simulation?'''),

                                              dcc.Checklist(id='yearsim', options=[
                                                    {'label': 'Annual simulation', 'value': 'yearsim'}],
                                                            value=[]
                                              ),
                                          ]),
                                          html.Div(style={'height': height_two}, children=[
                                              dcc.Markdown(style={'color': colors['black'],
                                                                  'borderBottom': '1px solid ' + str(colors['black']),
                                                                  'marginBottom': '1em'},
                                                           children='''SCOP?'''),

                                              dcc.Checklist(id='scop', options=[
                                                      {'label': 'SCOP normative', 'value': 'scop'}],
                                                            value=[]
                                              ),
                                          ])
                                      ]),

                             html.Div(className="four columns", children=[
                                 html.Div(id='sys-design-year',
                                          style={'height': height_two,
                                                 'borderLeft': '1px solid ' + str(colors['black']),
                                                 'padding': 10, 'display': 'none'},
                                          children=[
                                              dcc.Markdown(style={'color': colors['black'],
                                                                  'borderBottom': '1px solid ' + str(
                                                                      colors['black']),
                                                                  'marginBottom': '1em'},
                                                           children='''Which system design for annual Simulation?'''),

                                              dcc.Checklist(id='design-year', options=[
                                                      {'label': 'DIN', 'value': 'DIN'},
                                                      {'label': 'VDI', 'value': 'VDI'},
                                                      {'label': 'Individual design', 'value': 'ind'}
                                                  ],
                                                  value=['DIN', 'VDI']
                                              ),
                                          ]),

                                 html.Div(id='sys-design-scop',
                                          style={'height': height_two,
                                                 'borderLeft': '1px solid ' + str(colors['black']),
                                                 'padding': 10, 'display': 'none'},
                                          children=[
                                              dcc.Markdown(style={'color': colors['black'],
                                                                  'borderBottom': '1px solid ' + str(
                                                                      colors['black']),
                                                                  'marginBottom': '1em'},
                                                           children='''Which system design for SCOP calculation?'''),

                                              dcc.Checklist(id='design-scop', options=[
                                                      {'label': 'DIN', 'value': 'DIN'},
                                                      {'label': 'VDI', 'value': 'VDI'},
                                                      {'label': 'Individual design', 'value': 'ind'}
                                                  ],
                                                  value=['DIN', 'VDI']
                                              ),
                                          ]),
                             ]),

                             html.Div(id='ind-design', className="four columns",
                                      style={'borderLeft': '1px solid ' + str(colors['black']), 'padding': 10,
                                             'display': 'none'},
                                      children=[
                                          html.Div(style={'height': height_two, 'verticalAlign': 'middle'}, children=[
                                              html.Div(style={'width': '50%', 'display': 'inline-block',
                                                              'marginBottom': '5em'},
                                                       children=[
                                                           dcc.Markdown(style={'color': colors['black'],
                                                                               'borderBottom': '1px solid ' + str(
                                                                                   colors['black']),
                                                                               'marginBottom': '1em'},
                                                                        children='''Power heatpump in kW'''),

                                                           dcc.Input(id='p-hp',
                                                                     placeholder='Enter a value...',
                                                                     type='text',
                                                                     value=None
                                                                     )
                                                       ]),

                                              html.Div(style={'width': '50%', 'display': 'inline-block',
                                                              'marginBottom': '5em'},
                                                       children=[
                                                           dcc.Markdown(style={'color': colors['black'],
                                                                               'borderBottom': '1px solid ' + str(
                                                                                   colors['black']),
                                                                               'marginBottom': '1em'},
                                                                        children='''Power heatingrod in kW'''),

                                                           dcc.Input(id='p-hr',
                                                                     placeholder='Enter a value...',
                                                                     type='text',
                                                                     value=None
                                                                     )
                                                       ]),
                                          ]),

                                          html.Div(style={'height': height_two, 'verticalAlign': 'middle'}, children=[
                                              html.Div(style={'width': '50%', 'display': 'inline-block'}, children=[
                                                  dcc.Markdown(style={'color': colors['black'],
                                                                      'borderBottom': '1px solid ' + str(
                                                                          colors['black']),
                                                                      'marginBottom': '1em'},
                                                               children='''Volume hot water in l'''),

                                                  dcc.Input(id='vol-hw',
                                                            placeholder='Enter a value...',
                                                            type='text',
                                                            value=None
                                                            )
                                              ]),

                                              html.Div(style={'width': '50%', 'display': 'inline-block'}, children=[
                                                  dcc.Markdown(style={'color': colors['black'],
                                                                      'borderBottom': '1px solid ' + str(
                                                                          colors['black']),
                                                                      'marginBottom': '1em'},
                                                               children='''Volume domestic hot water in l'''),

                                                  dcc.Input(id='vol-dhw',
                                                            placeholder='Enter a value...',
                                                            type='text',
                                                            value=None
                                                            )
                                              ]),
                                          ]),
                                      ]),
                         ]),
            ]),  # end of first tab

            dcc.Tab(label='Building data', value='tab-building', children=[
                html.Div(className="row", style={'padding': 15, 'backgroundColor': colors['ebcgrey'],
                                                 'marginBottom': '2em', 'marginTop': '1em'}, children=[
                    html.Div(className="six columns", style={'padding': 10}, children=[
                        html.Div(children=[
                            dcc.Markdown(style={'color': colors['black'],
                                                'borderBottom': '1px solid ' + str(colors['black']),
                                                'marginBottom': '1em'},
                                         children='''How shall the demand be specified?'''),

                            dcc.RadioItems(id='demand-calc', options=[
                                    {'label': 'Teaser', 'value': 'teaser-demand'},
                                    {'label': 'Individuell', 'value': 'ind-demand'}
                                ],
                                value='teaser-demand'
                            ),
                        ]),
                    ]),
                    html.Div(className="six columns", style={'padding': 10}, children=[
                        html.Div(id='ind-temp', children=[
                            dcc.Markdown(style={'color': colors['black'],
                                                'borderBottom': '1px solid ' + str(colors['black']),
                                                'marginBottom': '1em'},
                                         children='''Heat demand at normative ambient temperature in kW'''),

                            dcc.Input(id='ind-demand',
                                      placeholder='Enter a value...',
                                      type='text',
                                      value=None
                                      )
                        ]),
                    ]),
                ]),
                dcc.Markdown(style={'color': colors['black'],
                                    'borderBottom': '1px solid ' + str(colors['black']),
                                    'marginBottom': '1em'},
                             children='''#### Building Data'''),
                html.Div(id='building-data', style={'padding': 15, 'backgroundColor': colors['ebcgrey'],
                                                    'marginBottom': '2em'}, children=[
                    html.Div(className="row", children=[
                        html.Div(className="six columns", style={'borderRight': '1px solid ' + str(colors['black']),
                                                                 },
                                 children=[
                            html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey'],
                                                             'marginBottom': '2em'}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Markdown(style={'color': colors['black'],
                                                        'marginBottom': '1em'},
                                                 children='''Building type'''),
                                ]),
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Dropdown(id='buildingtype', options=[
                                            {'label': 'Single family house', 'value': 'single_family_house'},
                                            {'label': 'Multi-family house', 'value': 'multi_family_house'},
                                            {'label': 'Apartment Block', 'value': 'apartment_block'},
                                            {'label': 'Terraced house', 'value': 'terraced_house'},
                                            {'label': 'Institute4', 'value': 'institute4'}
                                        ],
                                        value='single_family_house'
                                    ),
                                ]),
                            ]),
                            html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey'],
                                                             'marginBottom': '2em'}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Markdown(style={'color': colors['black'],
                                                        'marginBottom': '1em'},
                                                 children='''Construction year'''),
                                ]),
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Input(id='cons-year',
                                              placeholder='Enter a value...',
                                              type='text',
                                              value=None
                                              )
                                ]),
                            ]),
                            html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey'],
                                                             'marginBottom': '2em'}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Markdown(style={'color': colors['black'],
                                                        'marginBottom': '1em'},
                                                 children='''Number of floors'''),
                                ]),
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Input(id='n-floors',
                                              placeholder='Enter a value...',
                                              type='text',
                                              value=None
                                              )
                                ]),
                            ]),
                            html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey'],
                                                             'marginBottom': '2em'}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Markdown(style={'color': colors['black'],
                                                        'marginBottom': '1em'},
                                                 children='''Room height'''),
                                ]),
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Input(id='height',
                                              placeholder='Enter a value...',
                                              type='text',
                                              value=None
                                              )
                                ]),
                            ]),
                            html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey'],
                                                             'marginBottom': '2em'}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Markdown(style={'color': colors['black'],
                                                        'marginBottom': '1em'},
                                                 children='''Effective area'''),
                                ]),
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Input(id='area',
                                              placeholder='Enter a value...',
                                              type='text',
                                              value=None
                                              )
                                ]),
                            ]),
                        ]),
                        html.Div(id="location", className="six columns", style={}, children=[
                            html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey'],
                                                             'marginBottom': '2em'}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Markdown(style={'color': colors['black'],
                                                        'marginBottom': '1em'},
                                                 children='''Federal state'''),
                                ]),
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Dropdown(id='bundesland', options=[
                                            {'label': 'Baden Württemberg', 'value': 'Baden Württemberg'},
                                            {'label': 'Bayern', 'value': 'Bayern'},
                                            {'label': 'Berlin', 'value': 'Berlin'},
                                            {'label': 'Brandenburg', 'value': 'Brandenburg'},
                                            {'label': 'Bremen', 'value': 'Bremen'},
                                            {'label': 'Hamburg', 'value': 'Hamburg'},
                                            {'label': 'Hessen', 'value': 'Hessen'},
                                            {'label': 'Mecklenburg Vorpommern', 'value': 'Mecklenburg Vorpommern'},
                                            {'label': 'Niedersachsen', 'value': 'Niedersachsen'},
                                            {'label': 'Nordrhein Westfalen', 'value': 'Nordrhein Westfalen'},
                                            {'label': 'Rheinland Pfalz', 'value': 'Rheinland Pfalz'},
                                            {'label': 'Saarland', 'value': 'Saarland'},
                                            {'label': 'Sachsen', 'value': 'Sachsen'},
                                            {'label': 'Sachsen Anhalt', 'value': 'Sachsen Anhalt'},
                                            {'label': 'Schleswig Holstein', 'value': 'Schleswig Holstein'},
                                            {'label': 'Thüringen', 'value': 'Thüringen'},
                                        ],
                                        value='Baden Württemberg'
                                    ),
                                ]),
                            ]),
                            html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey'],
                                                             'marginBottom': '2em'}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Markdown(style={'color': colors['black'],
                                                        'marginBottom': '1em'},
                                                 children='''County'''),
                                ]),
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Dropdown(id='kreis', options=[
                                            {}
                                        ],
                                        value=None
                                    ),
                                ]),
                            ]),
                        ]),
                    ]),
                    dcc.Checklist(id='building-check', options=[
                            {'label': 'Modernised', 'value': 'mod'},
                            {'label': 'Ventilation unit', 'value': 'vent'}
                        ],
                        value=[]
                    ),
                ]),
            ]),
            # end of second tab
            dcc.Tab(label='common data', value='tab-common', children=[
                html.Div(style={'padding': 15, 'backgroundColor': colors['ebcgrey'],
                                'marginBottom': '2em'}, children=[
                    html.Div(className="row", children=[
                        html.Div(className="six columns", style={'borderRight': '1px solid ' + str(colors['black']),
                                                                 },
                                 children=[
                                     html.Div(className="row", style={'padding': 5,
                                                                      'backgroundColor': colors['ebcgrey'],
                                                                      'marginBottom': '2em'}, children=[
                                         html.Div(className="six columns", style={}, children=[
                                             dcc.Markdown(style={'color': colors['black'],
                                                                 'marginBottom': '1em'},
                                                          children='''DHW demand profile'''),
                                         ]),
                                         html.Div(className="six columns", style={}, children=[
                                             dcc.Dropdown(id='demand-profile', options=[
                                                     {'label': 'One person', 'value': 'einzelperson'},
                                                     {'label': 'Family with shower', 'value': 'familie_duschen'},
                                                     {'label': 'Family with shower and bathtub',
                                                      'value': 'familie_duschen_baden'},
                                                     # {'label': 'Individual profile', 'value': 'indprof'}
                                                 ],
                                                 value='einzelperson'
                                             ),
                                         ]),
                                     ]),
                                     html.Div(className="row", style={'padding': 5,
                                                                      'backgroundColor': colors['ebcgrey'],
                                                                      'marginBottom': '2em'}, children=[
                                         dcc.RadioItems(id='monovalent', options=[
                                                 {'label': 'Monovalent', 'value': 'mono'},
                                                 {'label': 'Bivalent', 'value': 'bi'}
                                             ],
                                             value='bi'
                                         ),
                                     ]),
                                     html.Div(className="row", style={'padding': 5,
                                                                      'backgroundColor': colors['ebcgrey'],
                                                                      'marginBottom': '2em'}, children=[
                                         html.Div(className="six columns", style={}, children=[
                                             dcc.Markdown(style={'color': colors['black'],
                                                                 'marginBottom': '1em'},
                                                          children='''Bivalent temperature'''),
                                         ]),
                                         html.Div(className="six columns", style={}, children=[
                                             dcc.Input(id='biv-temp',
                                                       placeholder='Enter a value...',
                                                       type='text',
                                                       value=None
                                                       )
                                         ]),
                                     ]),
                                     html.Div(className="row", style={'padding': 5,
                                                                      'backgroundColor': colors['ebcgrey'],
                                                                      'marginBottom': '2em'}, children=[
                                         html.Div(className="six columns", style={}, children=[
                                             dcc.Markdown(style={'color': colors['black'],
                                                                 'marginBottom': '1em'},
                                                          children='''Temperature of cold water'''),
                                         ]),
                                         html.Div(className="six columns", style={}, children=[
                                             dcc.Input(id='cold-water',
                                                       placeholder='Enter a value...',
                                                       type='text',
                                                       value=None
                                                       )
                                         ]),
                                     ]),
                                     html.Div(className="row", style={'padding': 5,
                                                                      'backgroundColor': colors['ebcgrey'],
                                                                      'marginBottom': '2em'}, children=[
                                         html.Div(className="six columns", style={}, children=[
                                             dcc.Markdown(style={'color': colors['black'],
                                                                 'marginBottom': '1em'},
                                                          children='''Temperature set value in buffer tank'''),
                                         ]),
                                         html.Div(className="six columns", style={}, children=[
                                             dcc.Input(id='tank',
                                                       placeholder='Enter a value...',
                                                       type='text',
                                                       value=None
                                                       )
                                         ]),
                                     ]),
                                 ]),
                        html.Div(className="six columns", style={}, children=[
                            dcc.Markdown(style={'color': colors['black'],
                                                'borderBottom': '1px solid ' + str(colors['black']),
                                                'marginBottom': '1em'},
                                         children='''#### Daily demands'''),

                            html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey'],
                                                             'marginBottom': '2em'}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Markdown(style={'color': colors['black'],
                                                        'marginBottom': '1em'},
                                                 children='''Daily volume demand'''),
                                ]),
                                html.Div(className="six columns", style={}, children=[
                                    html.Div(className="six columns", style={}, children=[
                                        dcc.Input(
                                            placeholder='Enter a value...',
                                            type='text',
                                            value=None
                                        )
                                    ]),
                                ]),
                            ]),
                            html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey'],
                                                             'marginBottom': '2em'}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Markdown(style={'color': colors['black'],
                                                        'marginBottom': '1em'},
                                                 children='''Daily heat demand'''),
                                ]),
                                html.Div(className="six columns", style={}, children=[
                                    html.Div(className="six columns", style={}, children=[
                                        dcc.Input(
                                            placeholder='Enter a value...',
                                            type='text',
                                            value=None
                                        )
                                    ]),
                                ]),
                            ]),
                            dcc.Markdown(style={'color': colors['black'],
                                                'borderBottom': '1px solid ' + str(colors['black']),
                                                'marginBottom': '1em'},
                                         children='''#### Critical period'''),

                            html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey'],
                                                             'marginBottom': '2em'}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Markdown(style={'color': colors['black'],
                                                        'marginBottom': '1em'},
                                                 children='''Duration of critical period'''),
                                ]),
                                html.Div(className="six columns", style={}, children=[
                                    html.Div(className="six columns", style={}, children=[
                                        dcc.Input(
                                            placeholder='Enter a value...',
                                            type='text',
                                            value=None
                                        )
                                    ]),
                                ]),
                            ]),
                            html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey'],
                                                             'marginBottom': '2em'}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Markdown(style={'color': colors['black'],
                                                        'marginBottom': '1em'},
                                                 children='''Heatloss during critical period'''),
                                ]),
                                html.Div(className="six columns", style={}, children=[
                                    html.Div(className="six columns", style={}, children=[
                                        dcc.Input(
                                            placeholder='Enter a value...',
                                            type='text',
                                            value=None
                                        )
                                    ]),
                                ]),
                            ]),
                            html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey'],
                                                             'marginBottom': '2em'}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Markdown(style={'color': colors['black'],
                                                        'marginBottom': '1em'},
                                                 children='''Heat demand during critical period'''),
                                ]),
                                html.Div(className="six columns", style={}, children=[
                                    html.Div(className="six columns", style={}, children=[
                                        dcc.Input(
                                            placeholder='Enter a value...',
                                            type='text',
                                            value=None
                                        )
                                    ]),
                                ]),
                            ]),
                        ]),
                    ]),
                ]),
            ]),
            # end of third tab
            dcc.Tab(label='DIN data', value='tab-din', children=[
                html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey']}, children=[
                    html.Div(className="six columns", style={}, children=[
                        html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey']}, children=[
                            html.Div(className="six columns", style={}, children=[
                                dcc.Markdown(style={'color': colors['black'],
                                                    'marginBottom': '1em'},
                                             children='''Design factor buffer tank'''),
                            ]),
                            html.Div(className="six columns", style={}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Input(id='design-fac-buffer',
                                              placeholder='Enter a value...',
                                              type='text',
                                              value=None
                                              )
                                ]),
                            ]),
                        ]),
                        html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey']}, children=[
                            html.Div(className="six columns", style={}, children=[
                                dcc.Markdown(style={'color': colors['black'],
                                                    'marginBottom': '1em'},
                                             children='''Design factor heatdemand'''),
                            ]),
                            html.Div(className="six columns", style={}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Input(id='design-fac-heat',
                                              placeholder='Enter a value...',
                                              type='text',
                                              value=None
                                              )
                                ]),
                            ]),
                        ]),
                        html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey']}, children=[
                            html.Div(className="six columns", style={}, children=[
                                dcc.Markdown(style={'color': colors['black'],
                                                    'marginBottom': '1em'},
                                             children='''Design factor hot water'''),
                            ]),
                            html.Div(className="six columns", style={}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Input(id='design-fac-water',
                                              placeholder='Enter a value...',
                                              type='text',
                                              value=None
                                              )
                                ]),
                            ]),
                        ]),
                    ]),
                    html.Div(className="six columns", style={}, children=[]),
                ]),
            ]),
            # end of fourth tab
            dcc.Tab(label='VDI data', value='tab-vdi', children=[
                html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey']}, children=[
                    html.Div(className="six columns", style={}, children=[
                        html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey']}, children=[
                            html.Div(className="six columns", style={}, children=[
                                dcc.Markdown(style={'color': colors['black'],
                                                    'marginBottom': '1em'},
                                             children='''Losses due to circulation'''),
                            ]),
                            html.Div(className="six columns", style={}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Input(id='circulation',
                                              placeholder='Enter a value...',
                                              type='text',
                                              value=None
                                              )
                                ]),
                            ]),
                        ]),
                        html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey']}, children=[
                            html.Div(className="six columns", style={}, children=[
                                dcc.Markdown(style={'color': colors['black'],
                                                    'marginBottom': '1em'},
                                             children='''Return temperature of the circulation'''),
                            ]),
                            html.Div(className="six columns", style={}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Input(id='circ-temp',
                                              placeholder='Enter a value...',
                                              type='text',
                                              value=None
                                              )
                                ]),
                            ]),
                        ]),
                        html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey']}, children=[
                            html.Div(className="six columns", style={}, children=[
                                dcc.Markdown(style={'color': colors['black'],
                                                    'marginBottom': '1em'},
                                             children='''Additional charge due to mixing losses in the storage tank'''),
                            ]),
                            html.Div(className="six columns", style={}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Input(id='fac-storage',
                                              placeholder='Enter a value...',
                                              type='text',
                                              value=None
                                              )
                                ]),
                            ]),
                        ]),
                        html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey']}, children=[
                            html.Div(className="six columns", style={}, children=[
                                dcc.Markdown(style={'color': colors['black'],
                                                    'marginBottom': '1em'},
                                             children='''Number of demand units (e.g. number of flats)'''),
                            ]),
                            html.Div(className="six columns", style={}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Input(id='num-units',
                                              placeholder='Enter a value...',
                                              type='text',
                                              value=None
                                              )
                                ]),
                            ]),
                        ]),
                        html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey']}, children=[
                            html.Div(className="six columns", style={}, children=[
                                dcc.Markdown(style={'color': colors['black'],
                                                    'marginBottom': '1em'},
                                             children='''Heat demand of other consumers (over 24h)'''),
                            ]),
                            html.Div(className="six columns", style={}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Input(id='dem-con',
                                              placeholder='Enter a value...',
                                              type='text',
                                              value=None
                                              )
                                ]),
                            ]),
                        ]),
                        html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey']}, children=[
                            html.Div(className="six columns", style={}, children=[
                                dcc.Markdown(style={'color': colors['black'],
                                                    'marginBottom': '1em'},
                                             children='''Blocked period (elctricity grid)'''),
                            ]),
                            html.Div(className="six columns", style={}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Input(id='bl-period',
                                              placeholder='Enter a value...',
                                              type='text',
                                              value=None
                                              )
                                ]),
                            ]),
                        ]),
                        html.Div(className="row", style={'padding': 5, 'backgroundColor': colors['ebcgrey']}, children=[
                            html.Div(className="six columns", style={}, children=[
                                dcc.Markdown(style={'color': colors['black'],
                                                    'marginBottom': '1em'},
                                             children='''Heating threshold'''),
                            ]),
                            html.Div(className="six columns", style={}, children=[
                                html.Div(className="six columns", style={}, children=[
                                    dcc.Input(id='heat-thresh',
                                              placeholder='Enter a value...',
                                              type='text',
                                              value=None
                                              )
                                ]),
                            ]),
                        ]),
                    ]),
                    html.Div(className="six columns", style={}, children=[]),
                ]),
            ])
        ]),
    ]),
    html.Div(children=[
        html.Button('Start calculation', id='start')
    ], style={'display': 'inline-block'}),
    html.Div(id='filler', style={'display': 'inline-block', 'width': '5%'}, children=[

    ]),
    html.Div(id='error', style={'display': 'inline-block'}, children=[

    ]),
    html.Div(id='run-message', style={'display': 'inline-block'}, children=[

    ]),

    html.Div(id='plots', style={'display': 'none'}, children=[

        # DIN results
        html.Div([
            dcc.Markdown(style={'color': colors['black'],
                                'borderBottom': '1px solid ' + str(colors['black']),
                                'marginBottom': '2em', 'marginTop': '2em'},
                         children='''#### Results DIN'''
                         ),
        ]),
        html.Div([
            dcc.Dropdown(
                id='filter-graph',
                options=[{'label': 'Heat Demand', 'value': 'demand_heat'},
                         {'label': 'Domestic hot water demand', 'value': 'demand_dhw'},
                         {'label': 'Produced heat of HP', 'value': 'produced_heat'},
                         {'label': 'Relative power of HP', 'value': 'relative_hp'},
                         {'label': 'Electrical demand of HP', 'value': 'electrical_hp'},
                         {'label': 'Electrical demand of HR', 'value': 'electrical_hr'},
                         {'label': 'Relative power of HR', 'value': 'relative_hr'},
                         {'label': 'COP', 'value': 'cop'},
                         {'label': 'Ambient temperature', 'value': 'amb_temp'},
                         {'label': 'Domestic hot water storage temperature at top', 'value': 'sto_dhw_top'},
                         {'label': 'Buffer storage temperature at top', 'value': 'sto_buf_top'}],
                value='demand_heat'
            )
        ], style={'width': '40%', 'padding': '0px 20px 20px 20px', 'marginTop': '5em'}),
        html.Div([
            dcc.Graph(style={'width': '90vh'}, id='graph',
                      # figure={'layout': {"height": '100%', 'width': '100%'}}
                      )
        ], style={'padding': '0 20'}),
        html.Div([
            dcc.Markdown(style={'color': colors['black'],
                                'borderBottom': '1px solid ' + str(colors['black']),
                                'marginBottom': '2em', 'marginTop': '2em'},
                         children='''From week:'''
                         ),
        ], style={'width': '100%', 'padding': '0 20'}),
        html.Div([
            dcc.Slider(
                id='time-slider-min',
                min=desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                    'DIN', 'result'))['time'].min()/168,
                max=desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                    'DIN', 'result'))['time'].max()/168,
                value=desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                      'DIN', 'result'))['time'].min()/168,
                marks={str(int(year/168)): str(int(year/168)) for year in
                       desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                       'DIN', 'result'))
                       [desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                        'DIN', 'result')).index % 168 == 0]['time']},
                step=None
            ),
         ], style={'width': '100%', 'padding': '0 20'}),
        html.Div([
            dcc.Markdown(style={'color': colors['black'],
                                'borderBottom': '1px solid ' + str(colors['black']),
                                'marginBottom': '2em', 'marginTop': '2em'},
                         children='''To week:'''),
        ], style={'width': '100%', 'padding': '0 20'}),
        html.Div([
            dcc.Slider(
                id='time-slider-max',
                min=desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                    'DIN', 'result'))['time'].min()/168,
                max=desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                    'DIN', 'result'))['time'].max()/168,
                value=desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                      'DIN', 'result'))['time'].max()/168,
                marks={str(int(year/168)): str(int(year/168)) for year in
                       desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results', 'DIN', 'result'))[
                           desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                           'DIN', 'result')).index % 168 == 0]['time']},
                step=None
            )
        ], style={'width': '100%', 'padding': '0 20'}),

        # VDI results

        html.Div([
            dcc.Markdown(style={'color': colors['black'],
                                'borderBottom': '1px solid ' + str(colors['black']),
                                'marginBottom': '2em', 'marginTop': '10em'},
                         children='''#### Results VDI'''
                         ),
        ]),
        html.Div([
            dcc.Dropdown(
                id='filter-graph-vdi',
                options=[{'label': 'Heat Demand', 'value': 'demand_heat'},
                         {'label': 'Domestic hot water demand', 'value': 'demand_dhw'},
                         {'label': 'Produced heat of HP', 'value': 'produced_heat'},
                         {'label': 'Relative power of HP', 'value': 'relative_hp'},
                         {'label': 'Electrical demand of HP', 'value': 'electrical_hp'},
                         {'label': 'Electrical demand of HR', 'value': 'electrical_hr'},
                         {'label': 'Relative power of HR', 'value': 'relative_hr'},
                         {'label': 'COP', 'value': 'cop'},
                         {'label': 'Ambient temperature', 'value': 'amb_temp'},
                         {'label': 'Domestic hot water storage temperature at top', 'value': 'sto_dhw_top'},
                         {'label': 'Buffer storage temperature at top', 'value': 'sto_buf_top'}],
                value='demand_heat'
            )
        ], style={'width': '40%', 'padding': '0px 20px 20px 20px', 'marginTop': '5em'}),
        html.Div([
            dcc.Graph(style={'width': '90vh'}, id='graph-vdi',
                      # figure={'layout': {"height": '100%', 'width': '100%'}}
                      )
        ], style={'padding': '0 20'}),
        html.Div([
            dcc.Markdown(style={'color': colors['black'],
                                'borderBottom': '1px solid ' + str(colors['black']),
                                'marginBottom': '2em', 'marginTop': '2em'},
                         children='''From week:'''
                         ),
        ], style={'width': '100%', 'padding': '0 20'}),
        html.Div([
            dcc.Slider(
                id='time-slider-min-vdi',
                min=desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                    'VDI', 'result'))['time'].min()/168,
                max=desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                    'VDI', 'result'))['time'].max()/168,
                value=desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                      'VDI', 'result'))['time'].min()/168,
                marks={str(int(year/168)): str(int(year/168)) for year in
                       desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                       'VDI', 'result'))
                       [desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                        'VDI', 'result')).index % 168 == 0]['time']},
                step=None
            ),
        ], style={'width': '100%', 'padding': '0 20'}),
        html.Div([
            dcc.Markdown(style={'color': colors['black'],
                                'borderBottom': '1px solid ' + str(colors['black']),
                                'marginBottom': '2em', 'marginTop': '2em'},
                         children='''To week:'''),
        ], style={'width': '100%', 'padding': '0 20'}),
        html.Div([
            dcc.Slider(
                id='time-slider-max-vdi',
                min=desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                    'VDI', 'result'))['time'].min()/168,
                max=desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                    'VDI', 'result'))['time'].max()/168,
                value=desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                      'VDI', 'result'))['time'].max()/168,
                marks={str(int(year/168)): str(int(year/168)) for year in
                       desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results', 'VDI', 'result'))[
                           desFunc.df_creator(os.path.join(Path(__file__).parents[0], 'Results',
                                                           'VDI', 'result')).index % 168 == 0]['time']},
                step=None
            )
        ], style={'width': '100%', 'padding': '0 20'}),
    ]),

    html.Div(id='trigger', children=0, style={'display': 'none'}),  # for Button disabling
    html.Div(id='dummy', style={'display': 'none'}),
    html.Div(id='dummy2', style={'display': 'none'}),
    html.Div(id='dummy3', style={'display': 'none'}),
    html.Div(id='dummy4', style={'display': 'none'}),
    html.Div(id='dummy5', style={'display': 'none'})
])


# TRY/individual value dialog is shown
@app.callback(
    Output('weather-dat', 'style'),
    Input('check-calc', 'value'))
def display_weather_option(calc):
    if 'heat-demand' in calc:
        return dict(**def_style, **{'display': 'block'})
    else:
        return dict(**def_style, **{'display': 'none'})


# dialog for individual temperature is shown
@app.callback(
    [Output('ind-weather', 'style'), Output('location', 'style')],
    [Input('check-calc', 'value'), Input('check-weather', 'value')])
def display_inidividual_weather(calc, weather):
    if 'heat-demand' in calc:
        if 'ind' in weather:
            return dict(**def_style, **{'display': 'block'}), dict(**box_style, **{'display': 'none'})
        else:
            return dict(**def_style, **{'display': 'none'}), dict(**box_style, **{'display': 'block'})
    else:
        return dict(**def_style, **{'display': 'none'}), dict(**box_style, **{'display': 'block'})


# dialog for system design for annual simulation
@app.callback(
    Output('sys-design-year', 'style'),
    Input('yearsim', 'value'))
def display_sys_design_annual(design_year):
    if 'yearsim' in design_year:
        return dict(**def_style, **{'display': 'block'})
    else:
        return dict(**def_style, **{'display': 'none'})


# reset values of dialog for system design for annual simulation
@app.callback(
    Output('design-year', 'value'),
    Input('yearsim', 'value'))
def reset_design_annual(design_year):
    return ['DIN', 'VDI']


# dialog for system design for scop
@app.callback(
    Output('sys-design-scop', 'style'),
    Input('scop', 'value'))
def display_sys_design_scop(design_scop):
    if 'scop' in design_scop:
        return dict(**def_style, **{'display': 'block'})
    else:
        return dict(**def_style, **{'display': 'none'})


# reset values of dialog for system design for scop
@app.callback(
    Output('design-scop', 'value'),
    Input('scop', 'value'))
def reset_design_scop(design_scop):
    return ['DIN', 'VDI']


# dialog for individual system design for annual simulation
@app.callback(
    Output('ind-design', 'style'),
    [Input('yearsim', 'value'), Input('scop', 'value'),
     Input('design-year', 'value'), Input('design-scop', 'value')])
def display_sys_design_ind(design_year, design_scop, design_year_ind, design_scop_ind):
    if 'yearsim' in design_year or 'scop' in design_scop:
        if 'ind' in design_year_ind or 'ind' in design_scop_ind:
            return dict(**def_style, **{'display': 'block'})
        else:
            return dict(**def_style, **{'display': 'none'})
    else:
        return dict(**def_style, **{'display': 'none'})


# update individual heat demand
@app.callback(
    [Output('ind-temp', 'style'), Output('building-data', 'style')],
    Input('demand-calc', 'value'))
def update_individual_demand(calc):
    if 'ind-demand' in calc:
        return dict({'display': 'block'}), dict(**box_style, **{'display': 'none'})
    else:
        return dict({'display': 'none'}), dict(**box_style, **{'display': 'block'})


# update bundesland drop down
@app.callback(
    Output('kreis', 'options'),
    Input('bundesland', 'value'))
def update_kreis_options(name):
    kreise = []
    wb = xl.load_workbook(os.path.join(dirpath, "Functions_GUI\\Wetterdaten\\Wetterzuordnung.xlsx"))
    seite1 = wb[str(name)]
    i = 1
    while i <= seite1.max_row:
        wert = seite1.cell(i, 2)
        if wert.value is not None:
            kreise.append(wert.value)
        i += 1
    return [{'label': m, 'value': m} for m in kreise]


# get data from calculation tab
@app.callback(
    Output('dummy', 'children'),
    [Input('check-calc', 'value'), Input('check-weather', 'value'),
     Input('ind-weatherval', 'value'), Input('scop', 'value'), Input('design-scop', 'value'), Input('yearsim', 'value'),
     Input('design-year', 'value'), Input('p-hp', 'value'), Input('p-hr', 'value'),
     Input('vol-hw', 'value'), Input('vol-dhw', 'value')])
def update_config_calculationtab(teaser, weather, weatherval, scop, scop_design, year, year_design, p_hp, p_hr,
                                 vol_hw, vol_dhw):
    if teaser == 'heat-demand':
        if weather == 'ind':
            if weather is not None and weather != '':
                data['T_heiz'] = float(weatherval)
            else:
                data['T_heiz'] = data_default['T_heiz']
    if 'scop' in scop:
        if 'ind' in scop_design:
            if p_hp is not None and p_hp != '':
                data['ind_dict']['Q_WP_AP'] = float(p_hp)
            else:
                data['ind_dict']['Q_WP_AP'] = data_default['ind_dict']['Q_WP_AP']
            if p_hr is not None and p_hr != '':
                data['ind_dict']['Q_HS_AP'] = float(p_hr)
            else:
                data['ind_dict']['Q_HS_AP'] = data_default['ind_dict']['Q_HS_AP']
            if vol_hw is not None and vol_hw != '':
                data['ind_dict']['V_Sp_WW'] = float(vol_hw)
            else:
                data['ind_dict']['V_Sp_WW'] = data_default['ind_dict']['V_Sp_WW']
            if vol_dhw is not None and vol_dhw != '':
                data['ind_dict']['V_Sp_TWW'] = float(vol_dhw)
            else:
                data['ind_dict']['V_Sp_TWW'] = data_default['ind_dict']['V_Sp_TWW']
    if 'yearsim' in year:
        if 'ind' in year_design:
            if p_hp is not None and p_hp != '':
                data['ind_dict']['Q_WP_AP'] = float(p_hp)
            else:
                data['ind_dict']['Q_WP_AP'] = data_default['ind_dict']['Q_WP_AP']
            if p_hr is not None and p_hr != '':
                data['ind_dict']['Q_HS_AP'] = float(p_hr)
            else:
                data['ind_dict']['Q_HS_AP'] = data_default['ind_dict']['Q_HS_AP']
            if vol_hw is not None and vol_hw != '':
                data['ind_dict']['V_Sp_WW'] = float(vol_hw)
            else:
                data['ind_dict']['V_Sp_WW'] = data_default['ind_dict']['V_Sp_WW']
            if vol_dhw is not None and vol_dhw != '':
                data['ind_dict']['V_Sp_TWW'] = float(vol_dhw)
            else:
                data['ind_dict']['V_Sp_TWW'] = data_default['ind_dict']['V_Sp_TWW']
    return 1  # just a dummy


# get data from building data tab
@app.callback(
    Output('dummy2', 'children'),
    [Input('demand-calc', 'value'), Input('buildingtype', 'value'),
     Input('cons-year', 'value'), Input('n-floors', 'value'), Input('height', 'value'), Input('area', 'value'),
     Input('building-check', 'value'), Input('bundesland', 'value'), Input('kreis', 'value'),
     Input('ind-demand', 'value')])
def update_config_buildingtab(teaser, buildingtype, consyear, n_floors, height_cei, area, check, bundesland, kreis,
                              q_ind):
    global fill_error
    if 'teaser-demand' in teaser:
        data['Gebaeudeart'] = buildingtype
        if "institute4" in buildingtype:
            data['residential'] = False
        if consyear is not None and consyear != '':
            data['Baujahr'] = int(consyear)
        else:
            data['Baujahr'] = data_default['Baujahr']
        if area is not None and area != '':
            data['Nutzflaeche'] = float(area)
        else:
            data['Nutzflaeche'] = data_default['Nutzflaeche']
        if height_cei is not None and height_cei != '':
            data['Deckenhoehe'] = float(height_cei)
        else:
            data['Deckenhoehe'] = data_default['Deckenhoehe']
        if n_floors is not None and n_floors != '':
            data['Geschosszahl'] = int(n_floors)
        else:
            data['Geschosszahl'] = data_default['Geschosszahl']
        if 'ven' in check:
            data['luefter'] = True
        else:
            data['luefter'] = False
        if 'mod' in check:
            if data['residential']:
                data['Modernisiert'] = "tabula_retrofit"
            else:
                data['Modernisiert'] = "heavy"
        else:
            if data['residential']:
                data['Modernisiert'] = "tabula_standard"
            else:
                data['Modernisiert'] = "light"
        data['bundesland'] = bundesland
        data['kreis'] = kreis
        wb = xl.load_workbook("Functions_GUI\\Wetterdaten\\Wetterzuordnung.xlsx")
        seite1 = wb[str(data['bundesland'])]
        i = 1
        while i <= seite1.max_row:
            wert = seite1.cell(i, 2)
            if wert.value == str(data['kreis']):
                wetterdaten = seite1.cell(i, 3).value + str('.mos')
                data['wetterdaten'] = wetterdaten
            i += 1
        if None in [buildingtype, consyear, n_floors, height_cei, area, check, bundesland, kreis]:
            fill_error = True
        else:
            fill_error = False
    elif 'ind-demand' in teaser:
        if q_ind is not None and q_ind != '':
            data['Q_ind'] = float(q_ind)
        else:
            data['Q_ind'] = data_default['Q_ind']
    return 1  # just a dummy


# get data from common data tab
@app.callback(
    Output('dummy3', 'children'),
    [Input('demand-profile', 'value'), Input('monovalent', 'value'),
     Input('biv-temp', 'value'), Input('cold-water', 'value'), Input('tank', 'value')])
def update_config_commontab(profile, monovalent, biv_temp, t_cold, tank_set):
    data['profil'] = profile
    if 'mono' in monovalent:
        data['bivalent'] = False
    elif 'bi' in monovalent:
        data['bivalent'] = True
    if biv_temp is not None and biv_temp != '':
        data['T_biv'] = float(biv_temp)
    else:
        data['T_biv'] = data_default['T_biv']
    if t_cold is not None and t_cold != '':
        data['T_kW'] = float(t_cold)
    else:
        data['T_kW'] = data_default['T_kW']
    if tank_set is not None and tank_set != '':
        data['T_TWW_set'] = float(tank_set)
    else:
        data['T_TWW_set'] = data_default['T_TWW_set']
    return 1  # just a dummy


# get data from din data tab
@app.callback(
    Output('dummy4', 'children'),
    [Input('design-fac-buffer', 'value'), Input('design-fac-heat', 'value'), Input('design-fac-water', 'value')])
def update_config_dintab(v_sp, f_hw, f_tww):
    if v_sp is not None and v_sp != '':
        data['v_Sp_WW'] = float(v_sp)
    else:
        data['v_Sp_WW'] = data_default['v_Sp_WW']
    if f_hw is not None and f_hw != '':
        data['f_HW'] = float(f_hw)
    else:
        data['f_HW'] = data_default['f_HW']
    if f_tww is not None and f_tww != '':
        data['f_TWW'] = float(f_tww)
    else:
        data['f_TWW'] = data_default['f_TWW']
    return 1  # just a dummy


# get data from vdi data tab
@app.callback(
    Output('dummy5', 'children'),
    [Input('circulation', 'value'), Input('circ-temp', 'value'), Input('fac-storage', 'value'),
     Input('num-units', 'value'), Input('dem-con', 'value'), Input('bl-period', 'value'),
     Input('heat-thresh', 'value')])
def update_config_vditab(circ, circ_temp, fac_storage, num_unit, dem_con, bl_period, heat_thresh):
    if circ is not None and circ != '':
        data['Q_zirk'] = float(circ)
    else:
        data['Q_zirk'] = data_default['Q_zirk']
    if circ_temp is not None and circ_temp != '':
        data['T_zirkRL'] = float(circ_temp)
    else:
        data['T_zirkRL'] = data_default['T_zirkRL']
    if fac_storage is not None and fac_storage != '':
        data['f_TWE'] = float(fac_storage)
    else:
        data['f_TWE'] = data_default['f_TWE']
    if num_unit is not None and num_unit != '':
        data['n_E'] = float(num_unit)
    else:
        data['n_E'] = data_default['n_E']
    if dem_con is not None and dem_con != '':
        data['Q_sonst'] = float(dem_con)
    else:
        data['Q_sonst'] = data_default['Q_sonst']
    if bl_period is not None and bl_period != '':
        data['t_SD'] = float(bl_period)
    else:
        data['t_SD'] = data_default['t_SD']
    if heat_thresh is not None and heat_thresh != '':
        data['T_HG'] = float(heat_thresh)
    else:
        data['T_HG'] = data_default['T_HG']
    return 1  # just a dummy


# disable button while calculation is performed
@app.callback(
    [Output('start', 'disabled'), Output('run-message', 'children')],
    [Input('start', 'n_clicks'), Input('trigger', 'children'), Input('error', 'children')])
def disable_button(clicked, trigger, error):
    context = [d['prop_id'] for d in dash.callback_context.triggered]
    print(context)
    context = dash.callback_context.triggered[0]['prop_id'].split('.')[0]
    # if the button triggered the function and button was pressed once already
    print(context)
    if 'start.n_clicks' in context and clicked is not None:
        if clicked > 0:
            return True, 'Calculation is running...'
    if 'trigger.children' in context:
        time.sleep(3)
        return False, ''
    else:
        return False, ''


# start calculation after button is pressed
@app.callback(
    [Output('error', 'children'), Output('trigger', 'children'), Output('plots', 'style')],
    [Input('start', 'n_clicks'), Input('demand-calc', 'value'), Input('check-calc', 'value'),
     Input('check-weather', 'value'), Input('yearsim', 'value'), Input('design-year', 'value'),
     Input('scop', 'value'), Input('design-scop', 'value'), Input('trigger', 'children')])
def start_calculation(clicked, teaser, calc, weather, year, design_year, scop, design_scop, trigger):
    global df_res_din
    global df_res_vdi
    global df_res_ind
    global click_counter
    if clicked is not None and clicked > click_counter:
        # counter, so that code is only running if button is clicked
        click_counter = click_counter + 1
        Func.write_config(data, abs_config_path)
        # data = Func.open_config(config_path)
        result_dict = {}
        if 'teaser-demand' in teaser:
            if not fill_error:
                if 'heat_demand' in calc:
                    if weather == 'try':
                        result_dict['heizlast'] = Func.calc_heizlast(abs_config_path)
                    if weather == 'ind':
                        result_dict['heizlast'] = Func.calc_heizlast(abs_config_path, ['T_heiz'])

                if 'DIN' in calc or 'VDI' in calc:
                    wetterdata_biv = Func.write_mos(data['wetterdaten'], data['T_biv'])
                    q_biv = Func.teasersim(abs_config_path, wetterdata_biv)[0]
                    wetterdata_na = Func.write_mos(data['wetterdaten'], data['T_NA'])
                    q_na = Func.teasersim(abs_config_path, wetterdata_na)[1]

                    if 'DIN' in calc:
                        result_dict['dict_din'] = Func.calc_din(abs_config_path, q_biv, q_na)
                    if 'VDI' in calc:
                        result_dict['dict_vdi'] = Func.calc_vdi(abs_config_path, q_biv, q_na)
            else:
                return 'An error occurred. Please check if you filled all data in \"Building Data\" tab.', 1, \
                       {'display': 'none'}

        else:
            if 'DIN' in calc:
                [q_biv, q_na] = Func.calc_normbiv(abs_config_path)
                result_dict['dict_din'] = Func.calc_din(abs_config_path, q_biv, q_na)
            if 'VDI' in calc:
                [q_biv, q_na] = Func.calc_normbiv(abs_config_path)
                result_dict['dict_vdi'] = Func.calc_vdi(abs_config_path, q_biv, q_na)

        if 'yearsim' in year:
            if 'DIN' in design_year or 'VDI' in design_year or 'ind' in design_year:
                if 'DIN' in design_year:
                    result_dict['path_jahres_din'] = Func.calc_js(abs_config_path, result_dict['dict_din'], 'DIN')
                    df_res_din = desFunc.df_creator(result_dict['path_jahres_din'])
                if 'VDI' in design_year:
                    result_dict['path_jahres_vdi'] = Func.calc_js(abs_config_path, result_dict['dict_vdi'], 'VDI')
                    df_res_vdi = desFunc.df_creator(result_dict['path_jahres_vdi'])
                if 'ind' in design_year:
                    result_dict['path_jahres_ind'] = Func.calc_js(abs_config_path, data['ind_dict'], 'Ind', ind=True)
                    df_res_ind = desFunc.df_creator(result_dict['path_jahres_ind'])
                yearplot = True
            else:
                return 'It was no Design selected. Therefore a annual simulation was not possible. ' \
                       'Please select a normative Design the next time.', 2, {'display': 'none'}
        else:
            yearplot = False

        if 'scop' in scop:
            if 'DIN' in design_scop or 'VDI' in design_scop or 'ind' in design_scop:
                if 'DIN' in design_scop:
                    result_dict['scop_din_points'] = Func.calc_points(abs_config_path, result_dict['dict_din'], 'DIN')
                    result_dict['scop_din'] = Func.scop_norm(abs_config_path, result_dict['scop_din_points'])
                if 'VDI' in design_scop:
                    result_dict['scop_vdi_points'] = Func.calc_points(abs_config_path, result_dict['dict_vdi'], 'VDI')
                    result_dict['scop_vdi'] = Func.scop_norm(abs_config_path, result_dict['scop_vdi_points'])
                if 'ind' in design_scop:
                    result_dict['scop_ind_points'] = Func.calc_points(abs_config_path, data['ind_dict'], 'IND',
                                                                      ind=True)
                    result_dict['scop_ind'] = Func.scop_norm(abs_config_path, result_dict['scop_ind_points'])
            else:
                return 'It was no normative Design selected. Therefore a calculation of the COP was not possible. ' \
                       'Please select a normative Design the next time.', 3, {'display': 'none'}
        print(result_dict)
        if yearplot:
            print(result_dict)
            return 'Calculation is Done. See the results below.', 4, {'display': 'block'}
        else:
            return 'Calculation is Done. See the results below.', 5, {'display': 'none'}
    return '', 0, {'display': 'none'}


def create_time_series(dff, axis_type, title):

    fig = px.scatter(dff, x='Year', y='Value')

    fig.update_traces(mode='lines+markers')

    fig.update_xaxes(showgrid=False)

    fig.update_yaxes(type='linear' if axis_type == 'Linear' else 'log')

    fig.add_annotation(x=0, y=0.85, xanchor='left', yanchor='bottom',
                       xref='paper', yref='paper', showarrow=False, align='left',
                       bgcolor='rgba(255, 255, 255, 0.5)', text=title)

    fig.update_layout(height=225, margin={'l': 20, 'b': 30, 'r': 10, 't': 10})

    return fig


@app.callback(
    dash.dependencies.Output('graph', 'figure'),
    [dash.dependencies.Input('filter-graph', 'value'), dash.dependencies.Input('time-slider-min', 'value'),
     dash.dependencies.Input('time-slider-max', 'value')])
def update_graph_din(y, time_slider_min, time_slider_max):
    global df_res_din
    time_slider_max = int(time_slider_max)
    time_slider_min = int(time_slider_min)
    df_plot_din = df_res_din[df_res_din['time']/168 >= time_slider_min]
    df_plot_din = df_plot_din[df_plot_din['time']/168 <= time_slider_max]
    fig = desFunc.visualize_res(df_plot_din, y)
    return fig


@app.callback(
    dash.dependencies.Output('graph-vdi', 'figure'),
    [dash.dependencies.Input('filter-graph-vdi', 'value'), dash.dependencies.Input('time-slider-min-vdi', 'value'),
     dash.dependencies.Input('time-slider-max-vdi', 'value')])
def update_graph_vdi(y, time_slider_min, time_slider_max):
    global df_res_vdi
    time_slider_max = int(time_slider_max)
    time_slider_min = int(time_slider_min)
    df_plot_vdi = df_res_vdi[df_res_vdi['time']/168 >= time_slider_min]
    df_plot_vdi = df_plot_vdi[df_plot_vdi['time']/168 <= time_slider_max]
    fig = desFunc.visualize_res(df_plot_vdi, y)
    return fig


if __name__ == '__main__':
    app.run_server(debug=True)
