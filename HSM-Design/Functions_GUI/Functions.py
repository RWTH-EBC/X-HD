import json
import numpy as np
import os
import sys
from teaser.project import Project
import DyMat
from scipy.io import savemat, loadmat
import openpyxl as xl
from pathlib import Path


# paths
PATH_DYMOLA = r"C:\Program Files\Dymola 2021"
Aixlib_path = r"C:\Users\aku-fst\AixLib_Teaser\AixLib\Aixlib"
Output_path = os.path.join(r"C:\Users\aku-fst\Teaser_Output")
comp_path = r"C:/Program Files (x86)/Microsoft Visual Studio 14.0/VC"
WPSmodel_path = r"D:\Remote-User\aku-fst\Hiwi\normativ-hps-design\Models\HPS_model\package.mo"  # path to HPS model
res_path = os.path.join(Path(__file__).parents[1], 'Results')
if not os.path.exists(os.path.join(PATH_DYMOLA, "Modelica", "Library", "python_interface")):
    sys.exit("[ERROR] Python Interface was not found in given folder. Program will be aborted!")
sys.path.insert(0, os.path.join(PATH_DYMOLA, r'Modelica\Library\python_interface\dymola.egg'))

if not os.path.exists(res_path):
    os.mkdir(res_path)
    print("Directory ", res_path,  " Created ")

from dymola.dymola_interface import DymolaInterface


def jahressim(model_path, model_name, result_path, initial_values,
              stoptime=32400000):
    """Simulates a simple heat pump system to get performance data over a year. model_path is the path to the
    corresponding modelica model. model_name is the name of the model in Dymola. result_path is the path in which
    the result file is stored. initial_values is a list with the design of the WPS in the order: [q_flow_hp_biv,
    q_flow_hr, v_buffer_sto, v_dhw_sto]. stoptime describes the simulation time. default is one year plus ten days
    for initialization. Returns the path to the result file."""
    dymola = DymolaInterface(dymolapath=os.path.join(PATH_DYMOLA, 'bin64', 'Dymola.exe'),
                             showwindow=True)
    try:

        dymola.openModel(path=os.path.join(Aixlib_path, 'package.mo'))
        dymola.openModel(model_path)
        dymola.SetDymolaCompiler("vs", ["CCompiler=MSVC", "MSVCDir=" + comp_path])
        dymola.experimentSetupOutput(events=False)
        # Parameters to change
        initial_names = ['Q_flow_hp_biv', 'Q_flow_hr', 'v_buffer_sto',
                         'v_dhw_sto']
        dymola.simulateExtendedModel(
            problem=model_name,
            startTime=0.0,
            stopTime=stoptime,
            outputInterval=3600,
            tolerance=0.001,
            resultFile=os.path.join(result_path, 'result'),
            initialNames=initial_names,
            initialValues=initial_values
        )
    except ValueError:
        print("[ERROR with Dymola. Please check model]")
    dymola.close()
    return os.path.join(result_path, 'result.mat')


def teasersim(config_name, weatherdatakreis=None, create_heat_mat=False):
    """Simulates Teaser model based on building inputs stored in the config, which path is given in config_name.
    weatherdatakreis is an optional argument to use a individual .mos file for the simulation. The default is the
    .mos file specified in the config file. create_heat_mat is Boolean, if it is set to true, a mat file with the
    heat demand over time is stored in the mat_files folder. Default is false. Returns a list with the heat demand
    averaged over one year and averaged over a cold month (january + febraury). The first is used to calculate the
    bivalent temperature with an edited .mos file for one year with the ambient temperature set constant to the
    bivalent temperature. The second is used for the normative temperature with the corresponding ambient temperature
    set constant over january and february. """
    data = open_config(config_name)

    if weatherdatakreis is None:
        weatherdatakreis = data['wetterdaten']

    # model/project name
    name_project = str(data['kreis'])
    name_model = str(data['Gebaeudeart']) + '_' + str(data['Baujahr'])

    w_path = os.getcwd()
    dymola_path = PATH_DYMOLA
    model_path = Output_path
    weatherdata = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Wetterdaten\\cop_app_mos')

    # start teaser
    prj = Project(load_data=True)
    prj.name = name_project
    if data['residential']:
        prj.add_residential(
            method='tabula_de',
            usage=data['Gebaeudeart'],
            name=name_model,
            year_of_construction=data['Baujahr'],
            number_of_floors=data['Geschosszahl'],
            height_of_floors=data['Deckenhoehe'],
            net_leased_area=data['Nutzflaeche'],
            with_ahu=data['luefter'],
            construction_type=data['Modernisiert']
        )
    else:
        prj.add_non_residential(
            method='bmvbs',
            usage=data['Gebaeudeart'],
            name=name_model,
            year_of_construction=data['Baujahr'],
            number_of_floors=data['Geschosszahl'],
            height_of_floors=data['Deckenhoehe'],
            net_leased_area=data['Nutzflaeche'],
            with_ahu=data['luefter'],
            construction_type=data['Modernisiert']
        )

    # exports model
    prj.used_library_calc = 'AixLib'
    prj.number_of_elements_calc = 2
    prj.weather_file_path = os.path.join(weatherdata, weatherdatakreis)
    prj.calc_all_buildings()
    prj.export_aixlib(
        # internal_id=None,
        path=Output_path)

    dymola = DymolaInterface(dymolapath=os.path.join(dymola_path, 'bin64', 'Dymola.exe'),
                             showwindow=True)
    try:
        dymola.openModel(path=os.path.join(Aixlib_path, 'package.mo'))
        dymola.experimentSetupOutput(events=False)

        # simulate with Dymola
        dymola.openModel(path=os.path.join(model_path, name_project, 'package.mo'))
        dymola.SetDymolaCompiler("vs", ["CCompiler=MSVC",
                                        "MSVCDir=" + comp_path])
        dymola.simulateModel(
            problem=name_project + '.' + name_model + '.' + name_model,
            startTime=0.0,
            stopTime=32400000,  # 1 year plus 10 days
            outputInterval=3600,
            method="Cvode",
            tolerance=0.001,
            resultFile=os.path.join(model_path, name_project, name_model, 'result'),
        )
    except ValueError:
        print("[ERROR]")

    # evaluation and analysis
    finally:
        dymola.close()

        os.chdir(w_path)
        jahr_auswertung = 240  # 2 days of initialization
        januar_auswertung = 240  # 2 days of initialization
        sum_jahr = 0  # counter for heat demand over 1 year
        sum_januar = 0  # counter for heat demand over 1 year (at normative ambient temperature)
        dmf = DyMat.DyMatFile(os.path.join(model_path, name_project, name_model, 'result'))
        dmf_d = dmf.data('multizone.PHeater[1]')  # variable for heat demand
        if create_heat_mat:
            mat_list = []
            time_list = np.arange(0, len(dmf_d) * 3600, 3600)
            for i in range(0, len(time_list)):
                mat_list.append([time_list[i], dmf_d[i]])
            mat = np.array(mat_list)
            path_heat = os.path.join(os.path.dirname(os.path.abspath(__file__)), r'matFiles\heat.mat')
            path_heat_dymola = os.path.join(Path(__file__).parents[1], 'Models', 'HPS_model', 'Components', 'new',
                                            'Data', 'heat.mat')
            savemat(path_heat, {'Q_dem': mat})
            savemat(path_heat_dymola, {'Q_dem': mat})
        while jahr_auswertung < dmf.size('multizone.PHeater[1]'):
            sum_jahr += dmf_d[jahr_auswertung]
            jahr_auswertung += 1
        while januar_auswertung < (744 + 240):
            sum_januar += dmf_d[januar_auswertung]
            januar_auswertung += 1
    heizlast = [sum_jahr / (8713 * 1000),
                sum_januar / (744 * 1000)]
    # print('The heat demand of the building is ' + str(heizlast))
    return heizlast


def write_mos(wetterdaten, temperature):
    """updates the TRY .mos file (path given as String in wetterdaten) and replaces the ambient temperature with the
    value given to the temperature (float) variable. Returns relative path to the new .mos file, which can be used for
    the weatherdatakreis argument from teasersim."""
    with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Wetterdaten\\cop_app_mos\\', wetterdaten),
              'r') as file:
        mos_list = file.readlines()

    header_list = mos_list[0:39]
    wetter_list = mos_list[39:]

    neuwetter_list = []
    for i in range(len(wetter_list)):
        neuwetter_list.append(wetter_list[i].split())
        neuwetter_list[i][1] = temperature
        neuwetter_list[i] = '\t'.join(map(str, neuwetter_list[i]))
    final_list = header_list + neuwetter_list

    with open(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           'Wetterdaten\\cop_app_mos\\erzeugt\\', wetterdaten), 'w') as f:
        for item in final_list[0:39]:
            f.write("%s" % item)
        for item in final_list[39:]:
            f.write("%s\n" % item)

    return 'erzeugt\\' + wetterdaten


def path_mat(config_name, ambtemp=True, dhw=True, heat=True, ind=False):
    """creates time series .mat files for the ambient temperature, heat demand and domestic hot water necessary for
    the HPS simulation and stores these at specific paths. ambtemp, dhw and heat are Boolean. If one of those is set
    to False, the corresponding .mat file is not created. Default for all is True. If ind (individual value for heat
    demand)is true, no teaser simulation is started"""
    data = open_config(config_name)
    # ambient temperature
    if ambtemp:
        with open(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               'Wetterdaten\\cop_app_mos\\', data['wetterdaten']), 'r') as file:
            mos_list = file.readlines()
        wetter_list = mos_list[39:]
        temp = [] * (len(wetter_list))
        time_list = np.arange(0, len(wetter_list) * 3600, 3600)
        ambtemp_list = []
        mat_list = []
        for i in range(0, len(wetter_list)):
            temp.append(wetter_list[i].split())
            ambtemp_list.append(float(temp[i][1]) + 273.15)  # TRY in °C, Simulation needs K
        for i in range(0, len(time_list)):
            mat_list.append([time_list[i], ambtemp_list[i]])
        mat = np.array(mat_list)
        path_temp = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'matFiles\\ambTemp.mat')
        path_temp_dymola = os.path.join(Path(__file__).parents[1], 'Models', 'HPS_model', 'Components', 'new',
                                        'Data', 'ambTemp.mat')
        savemat(path_temp, {'T_amb': mat})
        savemat(path_temp_dymola, {'T_amb': mat})
    # DHW
    if dhw:
        if data['profil'] == 'einzelperson':
            x = loadmat(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                     'matFiles\\profilesDHW\\Einzelperson.mat'))['DHW']
            savemat(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'matFiles\\DHW.mat'), {'DHW': x})
            savemat(os.path.join(Path(__file__).parents[1], 'Models', 'HPS_model', 'Components', 'new', 'Data',
                                 'DHW.mat'), {'DHW': x})
        if data['profil'] == 'familie_duschen':
            x = loadmat(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                     'matFiles\\profilesDHW\\Familie mit Duschen.mat'))['DHW']
            savemat(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'matFiles\\DHW.mat'), {'DHW': x})
            savemat(os.path.join(Path(__file__).parents[1], 'Models', 'HPS_model', 'Components', 'new', 'Data',
                                 'DHW.mat'), {'DHW': x})
        if data['profil'] == 'familie_duschen_baden':
            x = loadmat(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                     'matFiles\\profilesDHW\\Familie mit Duschen und Baden.mat'))['DHW']
            savemat(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'matFiles\\DHW.mat'), {'DHW': x})
            savemat(os.path.join(Path(__file__).parents[1], 'Models', 'HPS_model', 'Components', 'new', 'Data',
                                 'DHW.mat'), {'DHW': x})
    # heat demand
    if heat and not ind:
        dummy = teasersim(config_name, data['wetterdaten'], create_heat_mat=True)[1]
        # TODO: konstante mat file erzeugen


def open_config(config_name):
    """opens the given config file (path as string) and returns the data (dictionary) stored in it"""
    with open(config_name, 'r') as configfile:
        data = json.load(configfile)
    configfile.close()
    return data


def write_config(data, config_name):
    """writes data to the given config file (path as string). Data as dictionary."""
    with open(config_name, 'w') as configfile:
        json.dump(data, configfile)
    configfile.close()


def calc_heizlast(config_name, temperature=None):
    """uses Teaser to calculate the heat demand of the building over a year. A specific ambient temperature can be
    set to be present for the whole year. The default temperature is None, with which the temperature from the TRY is
    chosen."""
    data = open_config(config_name)
    if temperature is None:
        wetterdata = data['wetterdaten']
    else:
        wetterdata = write_mos(data['wetterdaten'], temperature)
    q_heiz = teasersim(config_name, wetterdata)[0]
    dict_heiz = {'gemittelte Heizlast über ein Jahr in kW': q_heiz}
    return dict_heiz


def calc_normbiv(config_name):
    """calculates the heat demand at the bivalent temperature using the normative method and returns a list with the
    heat demands at bivalent temperature and normative ambient temperature [q_hw_biv, q_hw_na]"""
    data = open_config(config_name)
    q_hw_na = data['Q_ind']
    q_hw_biv = 0
    if data['bivalent']:
        q_hw_biv = q_hw_na * (data['T_HG'] - data['T_biv']) / (data['T_HG'] - data['T_NA'])
    return [q_hw_biv, q_hw_na]


def calc_din(config_name, q_hw_biv, q_hw_na):
    """calculates the HPS design based on DIN 15450. q_hw_biv and q_hw_na are the heat demands at bivalent and norm
    temperature, which are returned by functions calc_norm_biv or teasersim. Returns a dictionary with the design"""
    data = open_config(config_name)

    # power of hp at normative ambient temperature
    q_wp_na = 0.46 * q_hw_na
    # volume dhw
    v_tww = data[data['profil']]['daily_vol']

    # heat demand at design point
    if data['bivalent']:
        q_hw_ap = q_hw_biv
    else:
        q_hw_ap = q_hw_na
    # storage dhw
    v_sp_tww = v_tww * (60 - data['T_kW']) / (data['T_TWW_set'] - data['T_kW'])
    # energy buffer storage
    q_s = data['c_w'] * (data['T_TWW_set'] - data['T_kW']) * v_tww
    # power dhw
    q_tww = (data[data['profil']]["Q_crit_DIN"] - q_s * (
            (data['T_TWW_set'] - 40) / (data['T_TWW_set'] - data['T_kW']))) / data[data['profil']][
                "t_crit_DIN"] + data[data['profil']]["Q_loss_crit"] / data[data['profil']]["t_crit_DIN"]
    # power of hp at designpoint
    if data['bivalent']:
        q_wp_ap = data['f_HW'] * q_hw_ap + data['f_TWW'] * q_tww
        q_hs_ap = q_hw_na - q_wp_na
    else:
        q_wp_ap = data['f_HW'] * q_hw_na + data['f_TWW'] * q_tww
        # power heating rod
        q_hs_ap = 0
    # buffer storage
    v_sp_ww_min = data['v_Sp_WW_min'] * q_wp_ap
    v_sp_ww_max = data['v_Sp_WW_max'] * q_wp_ap
    v_sp_ww = data['v_Sp_WW'] * q_wp_ap
    # output dict
    dict_din = {'V_Sp_TWW': v_sp_tww,
                'V_Sp_WW': v_sp_ww,
                'V_Sp_WW_min': v_sp_ww_min,
                'V_Sp_WW_max': v_sp_ww_max,
                'Q_WP_AP': q_wp_ap,
                'Q_HS_AP': q_hs_ap,
                'Q_HW_NA': q_hw_na}
    return dict_din


def calc_vdi(config_name, q_hw_biv, q_hw_na):
    """calculates the HPS design based on DIN 4645. q_hw_biv and q_hw_na are the heat demands at bivalent and norm
    temperature, which are returned by functions calc_norm_biv or teasersim. Returns a dictionary with the design"""
    data = open_config(config_name)

    # storage dhw
    v_zirk = data['Q_zirk'] / (data['c_w'] * (data['T_TWW_set'] - data['T_zirkRL']))
    q_dpb = data['n_E'] * data[data['profil']]["Q_crit_VDI"]
    v_dpb = q_dpb / (data['c_w'] * (data['T_TWW_set'] - data['T_kW']))
    v_sp_tww = (v_dpb + v_zirk) * data['f_TWE']
    # power dhw at design point
    q_tww_ap = data[data['profil']]['Q_crit_VDI'] / data[data['profil']]['t_crit_VDI']
    # power hp
    if data['bivalent']:
        q_wp_ap = (24 * q_hw_biv + data[data['profil']]["daily_Q"] + data['Q_sonst']) / (24 - data['t_SD'])
        q_hs_ap = q_hw_na + q_tww_ap - q_wp_ap
    else:
        q_wp_ap = (24 * q_hw_na + data[data['profil']]["daily_Q"] + data['Q_sonst']) / (24 - data['t_SD'])
        q_hs_ap = 0
    # buffer storage
    if data['bivalent']:
        v_sp_ww = 20 * q_wp_ap
    else:
        v_sp_ww = 35 * data['t_SD'] * q_wp_ap

    # output dict
    dict_vdi = {'V_Sp_TWW': v_sp_tww,
                'V_Sp_WW': v_sp_ww,
                'Q_WP_AP': q_wp_ap,
                'Q_HS_AP': q_hs_ap,
                'Q_HW_NA': q_hw_na}
    return dict_vdi


def calc_points(config_name, dict_norm, name, ind=False):
    """calculates different operation points of the HPS to determine the SCOP based on the normative method.
    dict_norm is a dict containing the dimensions of the HPS as it is returned by the functions calc_din and
    calc_vdi. name contains a string with the name of the result file of the simulation (recommended are either 'DIN'
    or 'VDI'). Returns a dict with the structure needed for the function scop_norm stored in SCOP_NORM."""
    data = open_config(config_name)
    tol = -15  # °C value of ambient temperature underneath which power is zero
    temperature = [tol, -10, -7, 2, 7, 12, data['T_biv']]  # °C
    temperature = sorted(temperature)

    # The WPS simulation requires a mat file for the heat demand, the DHW and the ambient temperature for all points,
    # that are required for the normative calculation of the SCOP
    cop_list = []
    maxheat_list = []
    # mat file for domestic hot water
    path_mat(config_name, ambtemp=False, heat=False, ind=ind)
    for i in temperature:
        # mat file for heat demand
        wetterdaten = write_mos(data['wetterdaten'], i)
        dummy = teasersim(config_name, wetterdaten, create_heat_mat=True)[1]

        # mat file for ambient temperature
        x = loadmat(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'matFiles\\ambTemp.mat'))['T_amb']
        x[:, 1] = i + 273.15
        savemat(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'matFiles\\ambTemp.mat'), {'T_amb': x})
        savemat(os.path.join(Path(__file__).parents[1], 'Models', 'HPS_model', 'Components', 'new', 'Data',
                             'ambTemp.mat'), {'T_amb': x})

        # simulation with beforehand calculated input data
        v_dhw_sto = dict_norm['V_Sp_TWW'] / 1000  # transformation of units
        v_buffer_sto = dict_norm['V_Sp_WW'] / 1000
        q_flow_hp_biv = dict_norm['Q_WP_AP'] * 1000
        q_flow_hr = dict_norm['Q_HS_AP'] * 1000

        # simulation of WPS
        initial_values = [q_flow_hp_biv, q_flow_hr, v_buffer_sto, v_dhw_sto]  # parameters changed in Simulation
        model_name = 'HPS_model.Components.new.SystemsFMU.System_MapOneHeatpump_v2'
        result_path = os.path.join(res_path, name)
        os.makedirs(result_path, exist_ok=True)
        sim_res_path = jahressim(WPSmodel_path, model_name, result_path, initial_values, stoptime=32400000)

        # evaluation of points
        res_points = DyMat.DyMatFile(sim_res_path)
        cop = res_points.data('generationCase.heatPump.division.u2')[240:]
        cop = np.mean(cop)
        print('der COP beträgt für ' + str(i) + '°C: ' + str(cop))
        cop_list.append(cop)
        heatpower_wp = res_points.data('generationCase.heatPump.SwitchHeatFlowCondenser.y')[240:]
        heatpower_wp = np.mean(heatpower_wp) / 1000
        maxheat_list.append(heatpower_wp)
        print('Die maximale Heizleistung der Wärmepumpe bei ' + str(i) + '°C beträgt: ' + str(heatpower_wp))
    punkt_list = ['Punkt_a', 'Punkt_b', 'Punkt_c', 'Punkt_d', 'Punkt_e', 'Punkt_f', 'Punkt_g']
    test = []
    for i in range(len(temperature)):
        if i == tol:
            test.append({'h_l': 0, 'cop': 0, 't': temperature[i]})
        else:
            test.append({'h_l': maxheat_list[i], 'cop': cop_list[i], 't': temperature[i]})
    punkte = dict(zip(punkt_list, test))
    return punkte


def calc_js(config_name, dict_norm, name, ind=False):
    """calculates the performance of a simple HPS system based on normative design. dict_norm is the dictionary
    containing the dimensions of the normative Designs, returned by calc_din and calc_vdi. name represents the name
    of the result .mat file of the HPS simulation (recommended are either 'DIN' or 'VDI')."""

    v_dhw_sto = dict_norm['V_Sp_TWW'] / 1000  # see modelica model and units
    v_buffer_sto = dict_norm['V_Sp_WW'] / 1000
    q_flow_hp_biv = dict_norm['Q_WP_AP'] * 1000
    q_flow_hr = dict_norm['Q_HS_AP'] * 1000
    path_mat(config_name, ind=ind)  # paths changed in Simulation
    initial_values = [q_flow_hp_biv, q_flow_hr, v_buffer_sto, v_dhw_sto]  # parameters changed in Simulation
    model_name = 'HPS_model.Components.new.SystemsFMU.System_MapOneHeatpump_v2'
    result_path = os.path.join(res_path, name)
    os.makedirs(result_path, exist_ok=True)
    sim_res_path = jahressim(WPSmodel_path, model_name, result_path, initial_values, stoptime=32400000)
    return sim_res_path


def scop_norm(config_name, points):
    data = open_config(config_name)
    # determines the ambient temperature and its variation throughout the year
    klima_moeglich = ["C", "A", "W"]
    klima = "W"
    if klima == "A":
        t_design = -10
    elif klima == "W":
        t_design = 2
    else:
        t_design = -22

    tol = -15  # °C value of ambient temperature underneath which power is zero
    c_dh = 0.9  # value that determines the influence of partial load on overall performance
    # assumed values from the example in NORM
    p_off = 0.0057  # kw power demand while turned off
    p_to = 0.0056  # kw power demand while temperature controll of
    p_sb = 0.0051  # kw power demand while standby
    p_ck = 0.0052  # kW power demand while crankcase heater turned on

    wetterdata_na = write_mos(data['wetterdaten'], data['T_NA'])
    p_designh = teasersim("Functions_GUI\\config\\config.json", wetterdata_na)[1]
    # Imports data from the norm saved in a excle file
    # Mainly the number of hours per year a specific ambient temperature is present
    # also number of hours per year the heat pump is in standby mode, on mode and so on
    wb = xl.load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)), "data_SCOPnorm\\AnhangDIN_EN_14825"
                                                                                   ".xlsx"))
    seite1 = wb["Tabelle1"]
    anhang = {}
    for zeile in range(3, seite1.max_row + 1):
        topkey = seite1.cell(zeile, 1)
        anhang[topkey.value] = {}
        for spalte in range(2, 5):
            key = seite1.cell(2, spalte)
            wert = seite1.cell(zeile, spalte)
            anhang[topkey.value][key.value] = wert.value
    zustandszeiten = {}
    for zeile in range(3, 8):
        topkey = seite1.cell(zeile, 6)
        zustandszeiten[topkey.value] = {}
        for spalte in range(7, 10):
            key = seite1.cell(2, spalte)
            wert = seite1.cell(zeile, spalte)
            zustandszeiten[topkey.value][key.value] = wert.value

    # calculates the heat demand for one Year based on the selected klima
    h_he = zustandszeiten["H_HE"][klima]
    q_h = p_designh * h_he

    # calculatesCR und COP_bin for data points
    # CR is the relation between heat demand and supply
    for x in points:
        pl = (points[x]["t"] - 16) / (t_design - 16)
        dc = points[x]["h_l"]
        cr = pl * p_designh / dc
        if cr > 1:
            cr = 1

        # Cop bin is the efficency including partial load losses
        cop_bin = points[x]["cop"] * cr / (c_dh * cr + (1 - c_dh))
        points[x]["Cop_Bin"] = cop_bin
    # sorting the performance points based on temperature for later interpolation
    sort = points.copy()
    reihenfolge = []
    temperatur = []
    z = 0
    i = len(sort)
    while i > 0:
        y = -300000
        for x in sort:
            if sort[x]["t"] > y:
                z = x
                y = sort[x]["t"]
        reihenfolge.append(z)
        temperatur.append(y)
        i -= 1
        del (sort[z])
    reihenfolge.reverse()
    temperatur.reverse()
    # Calculation of SCOP on and SCOP Net
    punkt_v = -1
    summe_nenner_scop_on = 0
    summe_zaehler_scop_on = 0
    summe_nenner_scop_net = 0
    summe_zaehler_scop_net = 0
    for i in range(-30, 16):  # looping throug all possible temperatures in the Norm
        h_j = anhang[i][klima]
        p_h = (i - 16) / (t_design - 16) * p_designh
        if i in temperatur:
            stelle = temperatur.index(i)
            q = points[reihenfolge[stelle]]["h_l"]
            cop_bin = points[reihenfolge[stelle]]["Cop_Bin"]
            punkt_v += 1
        else:  # interpolation
            if punkt_v >= 0:
                punkt = punkt_v
            else:
                punkt = 0
            if punkt >= len(reihenfolge) - 1:
                punkt = len(reihenfolge) - 2
            q_davor = points[reihenfolge[punkt]]["h_l"]
            q_danach = points[reihenfolge[punkt + 1]]["h_l"]

            t_davor = points[reihenfolge[punkt]]["t"]
            t_danach = points[reihenfolge[punkt + 1]]["t"]
            cop_bin_davor = points[reihenfolge[punkt]]["Cop_Bin"]
            cop_bin_danach = points[reihenfolge[punkt + 1]]["Cop_Bin"]
            if i < tol:
                q = 0
                cop_bin = 1
            else:
                q = ((q_danach - q_davor) / (t_danach - t_davor)) * (i - t_davor) + q_davor
                cop_bin = ((cop_bin_danach - cop_bin_davor) / (t_danach - t_davor)) * (i - t_davor) + cop_bin_davor
        if q > p_h:
            q = p_h
        if p_h > q:
            elbu = p_h - q
        else:
            elbu = 0

        nenner_scop_on = (((p_h - elbu) / cop_bin) + elbu) * h_j
        zaehler_scop_on = p_h * h_j
        zaehler_scop_net = h_j * (p_h - elbu)
        nenner_scop_net = ((p_h - elbu) / cop_bin) * h_j
        summe_nenner_scop_on += nenner_scop_on
        summe_zaehler_scop_on += zaehler_scop_on
        summe_zaehler_scop_net += zaehler_scop_net
        summe_nenner_scop_net += nenner_scop_net
    scop_on = summe_zaehler_scop_on / summe_nenner_scop_on
    scop_net = summe_zaehler_scop_net / summe_nenner_scop_net
    q_he = q_h / scop_on + zustandszeiten["H_TO"][klima] * p_to + zustandszeiten["H_SB"][klima] * p_sb + \
           zustandszeiten["H_CK"][klima] * p_ck + zustandszeiten["H_OFF"][klima] * p_off
    scop = q_h / q_he
    print("SCOP:", scop)
    return scop
