import sys
from PyQt5 import QtWidgets, uic
import openpyxl as xl
import Functions_GUI.Functions as Func
import os

# path to this module
dirpath = os.path.dirname(__file__)
# absolut path
abs_config_path = os.path.join(dirpath, 'Functions_GUI\\config\\config.json')

# Klasse für das Tool-Fenster zum Bearbeiten des Zapfprofils
class Profil(QtWidgets.QDialog):
    def __init__(self):
        super().__init__()
        # Laden des Tool-Designs
        uic.loadUi(os.path.join(dirpath, "Functions_GUI\\Tool_Profil.ui"), self)
        self.buttonBox.button(QtWidgets.QDialogButtonBox.Cancel).setText("Abbrechen und Eingaben löschen")
        # laden der configfile mit den default Werten
        self.data = Func.open_config(os.path.join(dirpath, 'Functions_GUI\\config\\default\\config_default.json'))
        # hier connections ür Tool-Window
        self.buttonBox.accepted.connect(self.profil)
        self.buttonBox.rejected.connect(self.clear)
        self.buttonBox.accepted.connect(lambda: Func.write_config(self.data, abs_config_path))

    def profil(self):
        """Funktion zum Auslesen der Daten aus dem Fenster Profil"""
        try:
            self.data['profil']['Q_loss_crit'] = float(self.QlossCrit_edit.text().replace(',', '.'))
            self.data['profil']['daily_vol'] = float(self.dailyVol_edit.text().replace(',', '.'))
            self.data['profil']['daily_Q'] = float(self.dailyQ_edit.text().replace(',', '.'))
            self.data['profil']['t_crit_DIN'] = float(self.tCritDIN_edit.text().replace(',', '.'))
            self.data['profil']['Q_crit_DIN'] = float(self.QCritDIN_edit.text().replace(',', '.'))
            self.data['profil']['t_crit_VDI'] = float(self.tCritVDI_edit.text().replace(',', '.'))
            self.data['profil']['Q_crit_VDI'] = float(self.QCritVDI_edit.text().replace(',', '.'))
        except ValueError:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setText("Bitte Angaben überprüfen")
            msg.setWindowTitle("Error")
            msg.exec_()

    def clear(self):
        """Funktion zum löschen der Daten aus dem Fenster Profil"""
        self.QlossCrit_edit.clear()
        self.dailyVol_edit.clear()
        self.dailyQ_edit.clear()
        self.tCritDIN_edit.clear()
        self.QCritDIN_edit.clear()
        self.tCritVDI_edit.clear()
        self.QCritVDI_edit.clear()


# Klasse für das Hauptfenster
class Ui(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        # Laden des GUI-Designs
        uic.loadUi(os.path.join(dirpath, "Functions_GUI\\UI_design.ui"), self)
        # Größe des Fensters anpassen
        self.setMinimumSize(1000, 500)
        # Variable zum Tracken eines Fehlers bei der Eingabe von Daten
        self.err = False
        # laden der configfile mit den default Werten
        self.data = Func.open_config(os.path.join(dirpath, 'Functions_GUI\\config\\default\\config_default.json'))
        # list with widgets of different tabs and correspondig keys for data☼
        self.widgetlist_gem = [self.bivalenztemp_edit, self.tempkalt_edit, self.tempSpeicher_edit]
        self.datalist_gem = ['T_biv', 'T_kW', 'T_TWW_set']
        self.widgetlist_din = [self.auslegungWSP_edit, self.auslegungsfakHL_edit, self.auslegungsfakWW_edit]
        self.datalist_din = ['v_Sp_WW', 'f_HW', 'f_TWW']
        self.widgetlist_vdi = [self.zirkverl_edit, self.zirkRLtemp_edit, self.aufschlagdurch_edit,
                               self.nutzungseinheiten_edit, self.sonstigerbedarf_edit, self.sperrdauer_edit,
                               self.heizgrenz_edit]
        self.datalist_vdi = ['Q_zirk', 'T_zirkRL', 'f_TWE', 'n_E',
                             'Q_sonst', 't_SD', 'T_HG']

        # hier connections einfügen
        self.heizlast_check.toggled.connect(self.showweather)
        self.indtemp_radio.toggled.connect(self.showtemp)
        self.jahres_check.toggled.connect(self.showjs)
        self.cop_check.toggled.connect(self.showcop)
        self.din_check.toggled.connect(self.showdin)
        self.vdi_check.toggled.connect(self.showvdi)
        self.indauslegung_check.toggled.connect(self.showind)
        self.indauslegungcop_check.toggled.connect(self.showind)
        self.teaser_radio.toggled.connect(self.showteaser)
        self.heizlast_radio.toggled.connect(self.showheat)
        self.monovalent_radio.toggled.connect(lambda x: self.bivalenztemp_edit.setDisabled(x))
        self.bundesland_combo.currentTextChanged.connect(self.bundeslandauswahl)
        # Workaround (not pretty) to change the current text of bundesland, so that the kreis combobox is updated
        self.bundesland_combo.setCurrentIndex(1)
        self.bundesland_combo.setCurrentIndex(0)
        self.tool_button.clicked.connect(self.profil_zeigen)
        self.dialog = Profil()
        self.start_button.clicked.connect(self.angabeberechnung)
        self.start_button.clicked.connect(self.heizlast)
        self.start_button.clicked.connect(self.gemeinsame_daten)
        self.start_button.clicked.connect(self.din)
        self.start_button.clicked.connect(self.vdi)
        self.start_button.clicked.connect(lambda: Func.write_config(self.data, abs_config_path))
        self.start_button.clicked.connect(self.berechnung)
        # Einstellen des Initialen Aussehens
        # tab berechnung
        self.din_check.setChecked(True)
        self.vdi_check.setChecked(True)
        self.try_radio.setChecked(True)
        self.indtemp_radio.setChecked(False)
        self.label_44.setHidden(True)
        self.try_radio.setHidden(True)
        self.indtemp_radio.setHidden(True)
        self.line_2.setHidden(True)
        self.label_43.setHidden(True)
        self.heiztemp_edit.setHidden(True)
        self.line_3.setHidden(True)
        # choices yearly simulation
        self.dinauslegung_check.setHidden(True)
        self.vdiauslegung_check.setHidden(True)
        self.indauslegung_check.setHidden(True)
        self.line_5.setHidden(True)
        # choices COP
        self.dinauslegungcop_check.setHidden(True)
        self.vdiauslegungcop_check.setHidden(True)
        self.indauslegungcop_check.setHidden(True)
        self.line_4.setHidden(True)
        # choices individual Design
        self.label_php.setHidden(True)
        self.powerhp_edit.setHidden(True)
        self.label_phr.setHidden(True)
        self.powerhr_edit.setHidden(True)
        self.label_ww.setHidden(True)
        self.volumeww_edit.setHidden(True)
        self.label_dhw.setHidden(True)
        self.volumedhw_edit.setHidden(True)
        self.line_6.setHidden(True)
        # tab gemeinsame daten
        self.teaser_radio.setChecked(True)
        self.widget.setHidden(True)
        self.bivalent_radio.setChecked(True)
        self.set_initial_design()

        self.show()

    def set_initial_design(self):
        """Set initial values and design for widgets"""
        # gemeinsame Daten
        for i, widget in enumerate(self.widgetlist_gem):
            widget.setPlaceholderText(str(self.data[self.datalist_gem[i]]))
        # DIN
        for i, widget in enumerate(self.widgetlist_din):
            widget.setPlaceholderText(str(self.data[self.datalist_din[i]]))
        # VDI
        for i, widget in enumerate(self.widgetlist_vdi):
            widget.setPlaceholderText(str(self.data[self.datalist_vdi[i]]))

    def profil_zeigen(self):
        """Zeigt Profil Fenster"""
        self.dialog.show()

    def angabeberechnung(self):
        """Funktion zum Auslesen der Daten aus dem Tab Berechnung"""
        if self.teaser_radio.isChecked():
            if self.indtemp_radio.isChecked():
                self.data['T_heiz'] = float(self.heiztemp_edit.text())
        if self.cop_check.isChecked():
            if self.indauslegung_check.isChecked():
                self.data['ind_dict']['Q_WP_AP'] = float(self.powerhp_edit.text())
                self.data['ind_dict']['Q_HS_AP'] = float(self.powerhr_edit.text())
                self.data['ind_dict']['V_Sp_WW'] = float(self.volumeww_edit.text())
                self.data['ind_dict']['V_Sp_TWW'] = float(self.volumedhw_edit.text())
        if self.jahres_check.isChecked():
            if self.indauslegungcop_check.isChecked():
                self.data['ind_dict']['Q_WP_AP'] = float(self.powerhp_edit.text())
                self.data['ind_dict']['Q_HS_AP'] = float(self.powerhr_edit.text())
                self.data['ind_dict']['V_Sp_WW'] = float(self.volumeww_edit.text())
                self.data['ind_dict']['V_Sp_TWW'] = float(self.volumedhw_edit.text())

    def heizlast(self):
        """Funktion zum Auslesen der Daten aus dem Tab Gebäudedaten"""
        if self.teaser_radio.isChecked():
            try:
                if self.nutzungsart_combo.currentText() == "Einfamilienhaus":
                    self.data['Gebaeudeart'] = "single_family_house"
                elif self.nutzungsart_combo.currentText() == "Reihenhaus":
                    self.data['Gebaeudeart'] = "terraced_house"
                elif self.nutzungsart_combo.currentText() == "Wohnblock":
                    self.data['Gebaeudeart'] = "apartment_block"
                elif self.nutzungsart_combo.currentText() == "Mehrfamilienhaus":
                    self.data['Gebaeudeart'] = "multi_family_house"
                elif self.nutzungsart_combo.currentText() == "Institut4":
                    self.data['Gebaeudeart'] = "institute4"
                    self.data['residential'] = False
                self.data['Baujahr'] = int(self.baujahr_edit.text())
                self.data['Nutzflaeche'] = float(self.nutzflaeche_edit.text())
                self.data['Deckenhoehe'] = float(self.raumhoehe_edit.text().replace(',', '.'))
                self.data['Geschosszahl'] = int(self.stockwerke_edit.text())
                self.data['luefter'] = bool(self.lueftung_check.isChecked())
                if self.modernisiert_check.isChecked():
                    if self.data['residential']:
                        self.data['Modernisiert'] = "tabula_retrofit"
                    else:
                        self.data['Modernisiert'] = "heavy"
                else:
                    if self.data['residential']:
                        self.data['Modernisiert'] = "tabula_standard"
                    else:
                        self.data['Modernisiert'] = "light"
                self.data['bundesland'] = self.bundesland_combo.currentText()
                self.data['kreis'] = self.kreis_combo.currentText()
                wb = xl.load_workbook(os.path.join(dirpath, "Functions_GUI\\Wetterdaten\\Wetterzuordnung.xlsx"))
                seite1 = wb[str(self.data['bundesland'])]
                i = 1
                while i <= seite1.max_row:
                    wert = seite1.cell(i, 2)
                    if wert.value == str(self.data['kreis']):
                        wetterdaten = seite1.cell(i, 3).value + str('.mos')
                        self.data['wetterdaten'] = wetterdaten
                    i += 1
                self.err = False
            except ValueError:
                self.err = True
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Warning)
                msg.setText("Bitte Angaben im Tab \"Heizlast\" überprüfen")
                msg.setWindowTitle("Error")
                msg.exec_()
        elif self.heizlast_radio.isChecked():
            try:
                self.data['Q_ind'] = float(self.heizlast_edit.text())
                self.err = False
            except ValueError:
                self.err = True
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Warning)
                msg.setText("Bitte Angaben im Tab \"Heizlast\" überprüfen")
                msg.setWindowTitle("Error")
                msg.exec_()

    def gemeinsame_daten(self):
        """Funktion zum Auslesen der Daten aus dem Tab gemeinsame Daten"""
        if self.zapf_combo.currentText() == "Einzelperson":
            self.data['profil'] = "einzelperson"
        elif self.zapf_combo.currentText() == "Familie mit Duschen":
            self.data['profil'] = "familie_duschen"
        elif self.zapf_combo.currentText() == "Familie mit Duschen und Baden":
            self.data['profil'] = "familie_duschen_baden"
        if self.monovalent_radio.isChecked():
            self.data['bivalent'] = False
        elif self.bivalent_radio.isChecked():
            self.data['bivalent'] = True
        for i, widget in enumerate(self.widgetlist_gem):
            if widget.text() == '':
                self.data[self.datalist_gem[i]] = float(widget.placeholderText())
            else:
                self.data[self.datalist_gem[i]] = float(widget.text())

    def showteaser(self, enabled):
        """radio button on tab Heizlast. If teaser is selected, building data is shown"""
        if enabled:
            self.groupBox.setHidden(False)
            self.widget.setHidden(True)

    def showheat(self, enabled):
        """radio button on tab Heizlast. If heizlast is selected, input box is shown"""
        if enabled:
            self.groupBox.setHidden(True)
            self.widget.setHidden(False)

    def showweather(self, enabled):
        """check box on tab Berechnung. If heizlast is selected, new buttons are shown"""
        if enabled:
            self.label_44.setHidden(False)
            self.try_radio.setHidden(False)
            self.indtemp_radio.setHidden(False)
            self.line_2.setHidden(False)
        else:
            self.label_44.setHidden(True)
            self.try_radio.setHidden(True)
            self.indtemp_radio.setHidden(True)
            self.line_2.setHidden(True)
            self.label_43.setHidden(True)
            self.heiztemp_edit.setHidden(True)
            self.line_3.setHidden(True)
            self.try_radio.setChecked(True)

    def showtemp(self, enabled):
        """radio button on tab Berechnung. If individuelle Temperatur is selected, input box is shown"""
        if enabled:
            self.label_43.setHidden(False)
            self.heiztemp_edit.setHidden(False)
            self.line_3.setHidden(False)
        else:
            self.label_43.setHidden(True)
            self.heiztemp_edit.setHidden(True)
            self.line_3.setHidden(True)

    def showjs(self, enabled):
        """check box on tab Berechnung. If jahressimulation is selected, new check boxes are shown"""
        if enabled:
            self.dinauslegung_check.setHidden(False)
            self.vdiauslegung_check.setHidden(False)
            self.indauslegung_check.setHidden(False)
            self.line_5.setHidden(False)
        else:
            self.dinauslegung_check.setHidden(True)
            self.vdiauslegung_check.setHidden(True)
            self.indauslegung_check.setHidden(True)
            self.indauslegung_check.setChecked(False)
            self.line_5.setHidden(True)

    def showcop(self, enabled):
        """check box on tab Berechnung. If COP nach Norm is selected, new check boxes are shown"""
        if enabled:
            self.dinauslegungcop_check.setHidden(False)
            self.vdiauslegungcop_check.setHidden(False)
            self.indauslegungcop_check.setHidden(False)
            self.line_4.setHidden(False)
        else:
            self.dinauslegungcop_check.setHidden(True)
            self.vdiauslegungcop_check.setHidden(True)
            self.indauslegungcop_check.setHidden(True)
            self.indauslegungcop_check.setChecked(False)
            self.line_4.setHidden(True)

    def showvdi(self, enabled):
        """check box on tab Berechnung. Depending on the choice, if the calculation according to VDI is selected or
        not, the options for the yearly simulation and the cop calculations are limited accordingly """
        if enabled:
            self.vdiauslegung_check.setChecked(True)
            self.vdiauslegung_check.setEnabled(True)
            self.vdiauslegungcop_check.setChecked(True)
            self.vdiauslegungcop_check.setEnabled(True)
        else:
            self.vdiauslegung_check.setChecked(False)
            self.vdiauslegung_check.setEnabled(False)
            self.vdiauslegungcop_check.setChecked(False)
            self.vdiauslegungcop_check.setEnabled(False)

    def showdin(self, enabled):
        """check box on tab Berechnung. If jahressimulation is selected, new check boxes are shown"""
        if enabled:
            self.dinauslegung_check.setChecked(True)
            self.dinauslegung_check.setEnabled(True)
            self.dinauslegungcop_check.setChecked(True)
            self.dinauslegungcop_check.setEnabled(True)
        else:
            self.dinauslegung_check.setChecked(False)
            self.dinauslegung_check.setEnabled(False)
            self.dinauslegungcop_check.setChecked(False)
            self.dinauslegungcop_check.setEnabled(False)

    def showind(self, enabled):
        """check box on tab Berechnung. If COP nach Norm is selected, new check boxes are shown"""
        if enabled:
            self.label_php.setHidden(False)
            self.powerhp_edit.setHidden(False)
            self.label_phr.setHidden(False)
            self.powerhr_edit.setHidden(False)
            self.label_ww.setHidden(False)
            self.volumeww_edit.setHidden(False)
            self.label_dhw.setHidden(False)
            self.volumedhw_edit.setHidden(False)
            self.line_6.setHidden(False)
        else:
            if not (self.indauslegungcop_check.isChecked() or self.indauslegung_check.isChecked()):
                self.label_php.setHidden(True)
                self.powerhp_edit.setHidden(True)
                self.label_phr.setHidden(True)
                self.powerhr_edit.setHidden(True)
                self.label_ww.setHidden(True)
                self.volumeww_edit.setHidden(True)
                self.label_dhw.setHidden(True)
                self.volumedhw_edit.setHidden(True)
                self.line_6.setHidden(True)

    def din(self):
        """Funktion zum Auslesen der Daten aus dem Tab DIN-Norm"""

        for i, widget in enumerate(self.widgetlist_din):
            if widget.text() == '':
                self.data[self.datalist_din[i]] = float(widget.placeholderText())
            else:
                self.data[self.datalist_din[i]] = float(widget.text())

    def vdi(self):
        """Funktion zum Auslesen der Daten aus dem Tab VDI-Norm"""

        for i, widget in enumerate(self.widgetlist_vdi):
            if widget.text() == '':
                self.data[self.datalist_vdi[i]] = float(widget.placeholderText())
            else:
                self.data[self.datalist_vdi[i]] = float(widget.text())

    def bundeslandauswahl(self):
        """Funktion zum Zuordnen der einzelnen Kreise aus den zugehörigen Bundesländern zur entsprechenden Combo-Box"""
        kreise = []
        wb = xl.load_workbook(os.path.join(dirpath, "Functions_GUI\\Wetterdaten\\Wetterzuordnung.xlsx"))
        self.data['bundesland'] = self.bundesland_combo.currentText()
        seite1 = wb[str(self.data['bundesland'])]
        i = 1
        self.kreis_combo.clear()
        while i <= seite1.max_row:
            wert = seite1.cell(i, 2)
            if wert.value is not None:
                kreise.append(wert.value)
            i += 1
        self.kreis_combo.setInsertPolicy(QtWidgets.QComboBox.InsertAlphabetically)
        self.kreis_combo.addItems(kreise)

    def berechnung(self):
        """Funktion zum Berechnen der Dimensionierung des WPS. Überprüft welche Berechnungen bereits durchgeführt
        wurden, um Doppelberechnungen zu umgehen"""
        if not self.err:
            if self.teaser_radio.isChecked():
                if self.heizlast_check.isChecked():
                    if self.try_radio.isChecked():
                        heizlast = Func.calc_heizlast(abs_config_path)
                        print('Die Heizlast beträgt: ' + str(heizlast))
                    if self.indtemp_radio.isChecked():
                        heizlast = Func.calc_heizlast(abs_config_path, self.data['T_heiz'])
                        print('Die Heizlast beträgt: ' + str(heizlast))

                if self.din_check.isChecked() or self.vdi_check.isChecked():
                    wetterdata_biv = Func.write_mos(self.data['wetterdaten'], self.data['T_biv'])
                    q_biv = Func.teasersim(abs_config_path, wetterdata_biv)[0]
                    wetterdata_na = Func.write_mos(self.data['wetterdaten'], self.data['T_NA'])
                    q_na = Func.teasersim(abs_config_path, wetterdata_na)[1]

                    if self.din_check.isChecked():
                        dict_din = Func.calc_din(abs_config_path, q_biv, q_na)
                        print('DIN: ' + str(dict_din) + '; Daten für Speicher in Liter, Leistungen in kW')
                    if self.vdi_check.isChecked():
                        dict_vdi = Func.calc_vdi(abs_config_path, q_biv, q_na)
                        print('VDI: ' + str(dict_vdi) + '; Daten für Speicher in Liter, Leistungen in kW')

            else:
                if self.din_check.isChecked():
                    print(Func.open_config(abs_config_path))
                    [q_biv, q_na] = Func.calc_normbiv(abs_config_path)
                    dict_din = Func.calc_din(abs_config_path, q_biv, q_na)
                    print('DIN: ' + str(dict_din) + '; Daten für Speicher in Liter, Leistungen in kW')
                if self.vdi_check.isChecked():
                    [q_biv, q_na] = Func.calc_normbiv(abs_config_path)
                    dict_vdi = Func.calc_vdi(abs_config_path, q_biv, q_na)
                    print('VDI: ' + str(dict_vdi) + '; Daten für Speicher in Liter, Leistungen in kW')

            if self.jahres_check.isChecked():
                if self.din_check.isChecked() or self.vdi_check.isChecked():
                    if self.dinauslegung_check.isChecked():
                        path_jahres = Func.calc_js(abs_config_path, dict_din, 'DIN')
                        print('Die Daten der Jahressimulation mit DIN Design liegen hier: ' + path_jahres)
                    if self.vdiauslegung_check.isChecked():
                        path_jahres = Func.calc_js(abs_config_path, dict_vdi, 'VDI')
                        print('Die Daten der Jahressimulation mit VDI Design liegen hier: ' + path_jahres)
                elif self.indauslegung_check.isChecked():
                    Func.calc_js(abs_config_path, self.data['ind_dict'], 'VDI', ind=True)
                else:
                    print('It was no normative Design selected. Therefore a calculation of the COP was not possible. '
                          'Please select a normative Design the next time.')

            if self.cop_check.isChecked():
                if self.din_check.isChecked() or self.vdi_check.isChecked():
                    if self.dinauslegungcop_check.isChecked():
                        Func.calc_points(abs_config_path, dict_din, 'DIN')
                    if self.vdiauslegungcop_check.isChecked():
                        Func.calc_points(abs_config_path, dict_vdi, 'VDI')
                else:
                    print('It was no normative Design selected. Therefore a calculation of the COP was not possible. '
                          'Please select a normative Design the next time.')
                if self.indauslegungcop_check.isChecked():
                    points = Func.calc_points(abs_config_path, self.data['ind_dict'], 'VDI', ind=True)
                    Func.scop_norm(abs_config_path, points)
            self.close()


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    w = Ui()
    sys.exit(app.exec_())
